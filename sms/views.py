import json
import os
import random
from pprint import pprint
import datetime
import time
import sys

from django.shortcuts import render


from django.db.models import Q
from datetime import timedelta
import xlsxwriter
from django.db.models import Sum
import calendar
from celery.result import AsyncResult
from celery.task import task
from django.db.models import OuterRef, Subquery, Sum
from django.contrib import messages
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.sites.shortcuts import get_current_site
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import connections
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, ListView
from django_chunked_iterator import iterator
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from rest_framework.response import Response
from mpesa_api.models import Mpesa
from roberms.celery import app
from sms.forms import *
from sms.models import *
from sms.tokens import account_activation_token
from sms.utils import *
from sms_api.views import send_usage
#from cgi import escape
import html
from html import escape
from openpyxl.styles import Border, Side

#html.escape(str).encode('ascii', 'xmlcharrefreplace')



def timed(f):
    @wraps(f)
    def wrapper(*args, **kwds):
        start = time()
        result = f(*args, **kwds)
        elapsed = time() - start
        print("%s took %d time to finish" % (f.__name__, elapsed))
        return result
    return wrapper


def is_user_customer(function):
    def wrap(request, *args, **kwargs):
        user = request.user
        customer = Customer.objects.filter(user_ptr_id=user.id).first()
        if customer is not None:
            return function(request, *args, **kwargs)
        elif CustomerSubAccounts.objects.filter(user_ptr_id=user.id).first():
            return function(request, *args, **kwargs)
        else:
            return redirect('sms:login')
    return wrap


class Merged:
    def __init__(self, phone_number, message):
        self.phone_number = phone_number
        self.message = message


# @login_required()
def home(request):
    return render(request, 'sms/sample_water.html')
def care_taker(request):
    return render(request, 'sms/caretaker.html')


def about(request):
    return render(request, 'sms/about.html')
def statement_date(request):
    return render(request, 'sms/statement_date.html')

@login_required()
@is_user_customer
def client_dashboard(request,client_id):
    #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    customer = WaterClientAll.objects.get(id=client_id)
    if customer is None:
        customer = WaterClientAll.objects.get(id=client_id)
        weeks = get_last_n_weeks(12)
        months = get_last_n_months(12)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        credit_usage = []
        for month in months:
            # print(week)
            messages = WaterMeterReadings.objects.filter(account_number=client_id,read_date__gte=one_month_ago, read_date__lte=current_day).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'months': months[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'phone_number': customer.msisdn,
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/client_apps.html', context)
    else:
        months = get_last_n_months(10)
        payments = WaterPaymentReceived.objects.filter(account_number=client_id)
        
        weeks = get_last_n_weeks(10)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        current_month = datetime.datetime.today()
        this_month = current_day.month
        
        the_sub_months=[]
        #the_sub_months=[2,1,12,11,10,9,8,7,6,5,4]
        #monthly_consumptions = [0, 10, 0, 5, 0, 0, 0, 0, 0, 0, 0, 15]
        monthly_consumptions = []
        for month in months:
        
    
            units_consumed_this = int(WaterMeterReadings.objects.filter(account_number=client_id,read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0)
            
            monthly_consumptions.append(units_consumed_this)
            the_sub_months.append(this_month)
            
            current_day = current_day - datetime.timedelta(days=30)
            this_month = current_day.month
            
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
            current_month = current_month - datetime.timedelta(days=30)
       # del the_sub_months[0]
          
        context = {
            'monthly_readings': monthly_consumptions[::-1],
            'weeks': weeks[::-1],
            'payments': payments,
            'months': the_sub_months[::-1],
            'phone_number': customer.msisdn,
            'client_id': customer.id,
            'amount_due': customer.amount_due,
            'client_name': customer.names,
            'client_court': customer.court,
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'courts': WaterCourt.objects.filter().count(),
            'readings': customer.last_meter_reading,
            'last_date': customer.last_meter_reading_date,
            
            'outbox': WaterMeterReadings.objects.filter().count(),
            'unallocated_payments': MiwamaMpesa.objects.filter(processed=2).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/client_apps.html', context)




@login_required()
@is_user_customer
def water_revenues(request):
    #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    customer = WaterClientAll.objects.all().first()
    if customer is None:
        customer = WaterClientAll.objects.all().first()
        weeks = get_last_n_weeks(12)
        months = get_last_n_months(12)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        credit_usage = []
        for month in months:
            # print(week)
            messages = WaterMeterReadings.objects.filter(account_number=client_id,read_date__gte=one_month_ago, read_date__lte=current_day).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'months': months[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'phone_number': customer.msisdn,
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/client_apps.html', context)
    else:
        months = get_last_n_months(10)
        payments = WaterPaymentReceived.objects.filter(account_number=1)
        
        weeks = get_last_n_weeks(10)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        current_month = datetime.datetime.today()
        this_month = current_day.month
        the_month_nam = current_day.month
        number_index = 1
        
        the_sub_months=[]
        the_month=[]
        #the_sub_months=[2,1,12,11,10,9,8,7,6,5,4]
        #monthly_consumptions = [0, 10, 0, 5, 0, 0, 0, 0, 0, 0, 0, 15]
        payments_r = []
        payments_e = []
        index_numbers = []
        mydeviation = []
        for month in months:
        
    
            payments_received = int(WaterPaymentReceived.objects.filter(pay_date__month=this_month).aggregate(total=Sum('amount'))['total'] or 0)
            expected_payments = int(WaterMeterReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('amount_from_units'))['total'] or 0)
            deviat = expected_payments - payments_received
            thiss_month = calendar.month_name[this_month]
            payments_r.append(payments_received)
            payments_e.append(expected_payments)
            index_numbers.append(number_index)
            mydeviation.append(deviat)
            the_sub_months.append(thiss_month)

            
            current_day = current_day - datetime.timedelta(days=30)
            this_month = current_day.month
            number_index = number_index + 1
            
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
            current_month = current_month - datetime.timedelta(days=30)
        receiving_months = the_sub_months[::1]
        myrevlist = zip(index_numbers,receiving_months, payments_e, payments_r, mydeviation)

       # del the_sub_months[0]

          
        context = {
            'revenues': payments_r[::1],
            'the_sub_months' : the_sub_months[::1],
            'revenue_list' : myrevlist,
            'weeks': weeks[::-1],
            'payments': payments,
            'months': the_sub_months[::1],
            'phone_number': customer.msisdn,
            'client_id': customer.id,
            'amount_due': customer.amount_due,
            'client_name': customer.names,
            'client_court': customer.court,
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'courts': WaterCourt.objects.filter().count(),
            'readings': customer.last_meter_reading,
            'last_date': customer.last_meter_reading_date,
            
            'outbox': WaterMeterReadings.objects.filter().count(),
            'unallocated_payments': MiwamaMpesa.objects.filter(processed=2).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/water_revenues.html', context)



def client_invoices(request, client_id):
    client = WaterClientAll.objects.get(id=client_id)
    invoices = WaterMeterReadings.objects.filter(account_number=client_id).order_by('-id').values()

    context = {
        'client': client,
        'invoices': invoices,
        'client_id': client_id
    }
    return render(request, 'sms/client_invoices.html', context)

def client_bills(request, client_id):
    client = WaterClientAll.objects.get(id=client_id)
    invoices = WaterMeterReadings.objects.filter(account_number=client_id).order_by('-id').values()

    context = {
        'client': client,
        'invoices': invoices
    }
    return render(request, 'sms/client_bills.html', context)

def create_invoice(request, client_id):
    client = Client.objects.get(id=client_id)

    if request.method == 'POST':
        invoice_number = ''
        last_invoice = Invoice.objects.all().order_by('id').last()
        if not last_invoice:
            invoice_number = 'RBLTD-100'
        else:
            cn = last_invoice.invoice_number
            cn_int = int(cn.split('RBLTD-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"RBLTD-{new_cn_int}"
            invoice_number = new_cn

        invoice = Invoice.objects.create(
            client=client,
            invoice_date=request.POST['invoice_date'],
            discount=request.POST['discount'],
            invoice_number=invoice_number,
            vat=int(request.POST['vat'])
        )
        messages.success(request, 'Invoice Creation Success');
        return redirect('Invoices:invoice_services', invoice.id)
    context = {
        'client': client
    }
    return render(request, 'invoices/create_invoice.html', context)
def invoice_preview(request, invoice_id):
    invoices = WaterMeterReadings.objects.filter(id=invoice_id)
    
    #client_s=WaterClientAll.objects.filter(id=invoices.meter_num)
    
    #services = Service.objects.filter(invoice=invoice)
    #client = WaterClientAll.objects.filter(id=13)
    total_amount = 0
    for service in invoices:
        total_amount += (float(service.amount_from_units))
        standing_charge = service.standing_charge
        rate = service.rate
        water_levy = service.water_levy
        invoice_number = service.id

        court = service.account_number.court
        network = service.account_number.network
        names = service.account_number.names
        tel = service.account_number.msisdn
        client_num = service.account_number.client_number
        
        units_consumed = service.units_consumed
        amount_from_units = service.amount_from_units
        invoice_account = service.account_number.names
        invoice_date = service.read_date
        current_month = datetime.datetime.today()
        read_month = invoice_date
        get_month = invoice_date.month
        get_date = invoice_date.day
        get_year= invoice_date.year
        gross_amount=units_consumed*rate
       # get_date=int(get_date)
        if get_date < 10:

            get_month=get_month-1
            if get_month==0:

                get_month=12
                get_year=get_year-1

        the_montis=calendar.month_name[get_month]
        
    discount = 0
  
        
        
    discount = 0
    
    new_total = total_amount - discount
    sub_total = new_total
    vat = new_total * 16/100
    #new_total += vat
    context = {
        'invoice': invoices,
        'services': invoices,
        'total': total_amount,
        'invoice_account':invoice_account,
        'new_total': new_total,
        'sub_total': sub_total,
        'discount': discount,
        'vat': vat,
        'rate': rate,
        'standing_charge': standing_charge,
        'water_levy': water_levy,
        'invoice_date': invoice_date,
        'invoice_number': invoice_number,
        'units_consumed': units_consumed,
        'amount_from_units': gross_amount,
        'the_montis': the_montis,
        'get_year': get_year,
        'get_date': get_date,
        'get_month': get_month,
        'names': names,
        'tel': tel,
        'network': network,
        'client_num': client_num,
        'courts': court,
        
    }
    return render(request, 'sms/invoice_preview.html', context)
def bill_preview(request, invoice_id):
    invoices = WaterMeterReadings.objects.filter(id=invoice_id)
    #services = Service.objects.filter(invoice=invoice)
    client = WaterClientAll.objects.filter(id=1)
    total_amount = 0
    for service in invoices:
        total_amount += (float(service.amount_from_units))
        standing_charge = service.standing_charge
        rate = service.rate
        water_levy = service.water_levy
        invoice_number = service.id
        units_consumed = service.units_consumed
        amount_from_units_read = service.amount_from_units
        amount_payable = service.payable
        brought_forward_arrears = service.arrears
        brought_forward_credit = service.credit
        invoice_date = service.read_date
        previous_readings = service.previous_reading
        current_readings = service.readings
        arrears = service.arrears
        credit = service.credit
        payent_account = service.account_number.id

        court = service.account_number.court
        network = service.account_number.network
        names = service.account_number.names
        tel = service.account_number.msisdn
        client_num = service.account_number.client_number




        if brought_forward_arrears>0:
            brought_forward = brought_forward_arrears
        else:
            brought_forward=brought_forward_credit*-1
        
        current_month = datetime.datetime.today()
        read_month = invoice_date
        get_month = invoice_date.month
        get_date = invoice_date.day
        get_year= invoice_date.year
        gross_amount=units_consumed*rate
       # get_date=int(get_date)
        if get_date < 10:

            get_month=get_month-1
            if get_month==0:

                get_month=12
                get_year=get_year-1

        the_montis=calendar.month_name[get_month]
        
    discount = 0
   
        
        
    discount = 0
    
    new_total = total_amount - discount
    sub_total = new_total
    vat = new_total * 16/100
    #new_total += vat
    context = {
        'invoice': invoices,
        'services': invoices,
        'total': total_amount,
        'new_total': new_total,
        'sub_total': sub_total,
        'discount': discount,
        'vat': vat,
        'rate': rate,
        'standing_charge': standing_charge,
        'water_levy': water_levy,
        'invoice_date': invoice_date,
        'invoice_number': invoice_number,
        'units_consumed': units_consumed,
        'amount_from_units': gross_amount,
        'amount_from_units_read': amount_from_units_read,
        'brought_forward_arrears': brought_forward_arrears,
        'brought_forward': brought_forward,
        'brought_forward_credit': brought_forward_credit,
        'the_montis': the_montis,
        'get_year': get_year,
        'get_date': get_date,
        'get_month': get_month,
        'names': names,
        'amount_payables': amount_payable,
        'tel': tel,
        'network': network,
        'client_num': client_num,
        'previous_readings': previous_readings,
        'current_readings': current_readings,
        'payent_account': payent_account,
        'amount_payable': total_amount,
        'credit': credit,
        'arrears': arrears,
        'courts': court,
        'client': client
    }
    return render(request, 'sms/bills_preview.html', context)
def statement_preview(request, client_id): 
    
    client = WaterClientAll.objects.filter(id=client_id)
    balance_bought_forward=0
    for clientn in client:            
            court = clientn.court
            network = clientn.network
            names = clientn.names
            tel = clientn.msisdn
            client_num = clientn.client_number      
    if request.method == 'POST':
        statement = WaterStatement.objects.filter(id=0)

       
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        starting_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        ending_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        yesteday_is = datetime.datetime.today() + datetime.timedelta(days=-1)
        tomorrow_is = datetime.datetime.today() + datetime.timedelta(days=1)

        plus_one_day = ending_date + datetime.timedelta(days=1)
        client = WaterClientAll.objects.filter(id=client_id).first()
        statement = WaterStatement.objects.filter(account_number=client.id,statement_date__range=[starting_date, plus_one_day])
        statement_first = WaterStatement.objects.filter(account_number=client.id,statement_date__range=[starting_date, plus_one_day]).first()
        balance_brought_forward=0
        stat_date = request.POST['start_date']  
        if statement_first:


            stat_date=statement_first.statement_date
            
            if statement_first.debit>0:
                balance_brought_forward=statement_first.balance-statement_first.debit
                
            else:
                balance_brought_forward=statement_first.balance+statement_first.credits

        invoice_date = client.created_at
        current_month = datetime.datetime.today()
        statement_day = datetime.datetime.today()
        read_month = invoice_date
        get_month = invoice_date.month
        get_date = invoice_date.day
        get_year= invoice_date.year
       
        if get_date < 10:

            get_month=get_month-1
            if get_month==0:

                get_month=12
                get_year=get_year-1

        the_montis=calendar.month_name[get_month]

        
        
        
    
    #new_total += vat
        context = {        
            'statement': statement,        
            'starting_date': starting_date,
            'ending_date': ending_date,
            'names': names,
            'tel': tel,
            'stat_date': stat_date,
            'balance_brought_forward': balance_brought_forward,
            'statement_day': statement_day,
            'network': network,
            'client_num': client_num,
            'client_id': client_id,
            'courts': court,
            'client': client
        }
        return render(request, 'sms/statement_preview.html', context)
    else:
        statement = WaterStatement.objects.filter(id=0)
        context = {        
            'statement': statement,
                    
            
            'names': names,
            'tel': tel,
            'client_id': client_id,
            
            'network': network,
            'client_num': client_num,
            'courts': court,
            'client': client
        }
        return render(request, 'sms/statement_preview.html', context)
        

def accountss_preview(request): 
    
    client = WaterMainReadings.objects.filter(id=1)
    balance_bought_forward=0
          
    if request.method == 'POST':
        statement = WaterStatement.objects.filter(id=0)

       
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        starting_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        ending_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        yesteday_is = datetime.datetime.today() + datetime.timedelta(days=-1)
        tomorrow_is = datetime.datetime.today() + datetime.timedelta(days=1)

        plus_one_day = ending_date + datetime.timedelta(days=1)
        client = WaterClientAll.objects.filter(id=client_id).first()
        statement = WaterStatement.objects.filter(account_number=client.id,statement_date__range=[starting_date, plus_one_day])
        statement_first = WaterStatement.objects.filter(account_number=client.id,statement_date__range=[starting_date, plus_one_day]).first()
        balance_brought_forward=0
        stat_date = request.POST['start_date']  
        if statement_first:


            stat_date=statement_first.statement_date
            
            if statement_first.debit>0:
                balance_brought_forward=statement_first.balance-statement_first.debit
                
            else:
                balance_brought_forward=statement_first.balance+statement_first.credits

        invoice_date = client.created_at
        current_month = datetime.datetime.today()
        statement_day = datetime.datetime.today()
        read_month = invoice_date
        get_month = invoice_date.month
        get_date = invoice_date.day
        get_year= invoice_date.year
       
        if get_date < 10:

            get_month=get_month-1
            if get_month==0:

                get_month=12
                get_year=get_year-1

        the_montis=calendar.month_name[get_month]

        
        
        
    
    #new_total += vat
        context = {        
            'statement': statement,        
            'starting_date': starting_date,
            'ending_date': ending_date,
            'names': names,
            'tel': tel,
            'stat_date': stat_date,
            'balance_brought_forward': balance_brought_forward,
            'statement_day': statement_day,
            'network': network,
            'client_num': client_num,
            'client_id': client_id,
            'courts': court,
            'client': client
        }
        return render(request, 'sms/statement_preview.html', context)
    else:
        statement = WaterStatement.objects.filter(id=0)
        context = {        
            'statement': statement,
                    
            
            'names': names,
            'tel': tel,
            'client_id': client_id,
            
            'network': network,
            'client_num': client_num,
            'courts': court,
            'client': client
        }
        return render(request, 'sms/statement_preview.html', context)
        


def multi_search(request): 
    
    client = WaterClientAll.objects.filter(id=0)
    for clientn in client:            
            court = clientn.court
            network = clientn.network
            names = clientn.names
            tel = clientn.msisdn
            client_num = clientn.client_number      
    if request.method == 'POST':
        statement = WaterStatement.objects.filter(id=0)

       
        search_key = request.POST['search_key']
        search_keys = request.POST['search_key']
        lname = 'BENARD BARONGO MOSOMI'
        
        my_binary_data = bytes(search_key, 'utf-8')
        yesteday_is = datetime.datetime.today() + datetime.timedelta(days=-1)
        tomorrow_is = datetime.datetime.today() + datetime.timedelta(days=1)
        
        

        

        if type(search_key)==str:
            client = WaterClientAll.objects.raw('SELECT * FROM sms_waterclientall where names=%s',[search_key])
            try:
                int(search_key)
                client = WaterClientAll.objects.raw('SELECT * FROM sms_waterclientall where id= %s or msisdn=%s or id_num=%s ', [search_key, search_key, search_key])
                payments = WaterPaymentReceived.objects.raw('SELECT * FROM sms_waterpaymentreceived where account_number= %s or dest_msisdn=%s', [search_key, search_key])
                readings = WaterMeterReadings.objects.raw('SELECT * FROM sms_watermeterreadings where account_number_id= %s or msisdn=%s', [search_key, search_key])
                bill_sent = WaterBillSent.objects.raw('SELECT * FROM sms_waterbillsent where account_number_id= %s or dest_msisdn=%s', [search_key, search_key])
                messages_sent = WaterOutbox.objects.raw('SELECT * FROM water_outbox where dest_msisdn=%s', [search_key])
            except ValueError:
                if len(search_key)>2:
                    client = WaterClientAll.objects.filter(names__icontains=search_key)
                if len(search_key)>3:
                    client = WaterClientAll.objects.filter(names__icontains=search_key) | WaterClientAll.objects.filter(provided_account__icontains=search_key)
                     
                
                #client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where id=2')
                #payments = WaterPaymentReceived.objects.raw('SELECT * FROM green_note.sms_waterpaymentreceived where account_number= %s or dest_msisdn=%s', [search_key, search_key])
                payments = WaterPaymentReceived.objects.filter(confirmation_code__exact=search_key)
                readings = WaterMeterReadings.objects.raw('SELECT * FROM sms_watermeterreadings where account_number_id= %s or msisdn=%s', [search_key, search_key])
                bill_sent = WaterBillSent.objects.raw('SELECT * FROM sms_waterbillsent where account_number_id= %s or dest_msisdn=%s', [search_key, search_key])
                messages_sent = WaterOutbox.objects.raw('SELECT * FROM water_outbox where dest_msisdn=%s', [search_key])




          #  client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where id= %s or msisdn=%s or id_num=%s', [search_key, search_key, search_key])
            #client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where id= 2')
          #  statement = WaterStatement.objects.filter(account_number=search_key)
        #elif type(search_key)==str:
        #    client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where names= %s',[search_key])
       # else:
            
        #    try:
            #    int(search_key)
             #   client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where id= %s or msisdn=%s or id_num=%s', [search_key, search_key, search_key])
           # except ValueError:
            #    client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where names= %s', [search_key])


            

        #client = WaterClientAll.objects.filter(Q(id=search_key))
        

        
        current_month = datetime.datetime.today()
        #client = WaterClientAll.objects.raw('SELECT * FROM green_note.sms_waterclientall where names LIKE %')
        statement_day = datetime.datetime.today()
        
       
        
        
        
    
    #new_total += vat
        context = {        
            'statement': statement,            
            'search_key': search_key,
            'readings': readings,
            'bills_sent': bills_sent,
            'messages_sent': messages_sent,
            'payments': payments,                        
            'profile': client
        }
        return render(request, 'sms/multi_search.html', context)
    else:
        statement = WaterStatement.objects.filter(id=0)
        context = {        
            'statement': statement,        
            
            
        }
        return render(request, 'sms/multi_search.html', context)
        


def main_meterchart(request):
  production_data_2014 = [
    { "y": 105.48, "label": "Rice"},
    { "y": 86.53, "label": "Wheat"},
    { "y": 42.86, "label": "Coarse Cereals"},
    { "y": 17.15, "label": "Total Pulses"}
  ]
  production_data_2015 = [
    { "y": 104.41, "label": "Rice"},
    { "y": 92.29, "label": "Wheat"},
    { "y": 38.52, "label": "Coarse Cereals"},
    { "y": 16.35, "label": "Total Pulses"}
  ]
  production_data_2016 = [
    { "y": 108.86, "label": "Rice"},
    { "y": 96.64, "label": "Wheat"},
    { "y": 44.34, "label": "Coarse Cereals"},
    { "y": 22.14, "label": "Total Pulses"}
  ]
  #return render(request, 'sms/main_meter.html', context)
  return render(request, 'sms/main_meterchart.html', { "production_data_2014" : production_data_2014, "production_data_2015": production_data_2015, "production_data_2016": production_data_2016 })

@login_required()
@is_user_customer
def main_meter(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is None:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        #weeks = get_last_n_weeks(12)
        months = get_last_n_months(7)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        this_month = current_day.month
        monthly_all_consumptions_main = []
        monthly_all_consumptions_clients = []
        monthly_all_consumptions_clientsw = [12,3,4,5,5,2,6]
        for month in months:
            # print(week)
            monthly_units = WaterMainReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0
            monthly_units_clients = WaterMeterReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0
               
            messages = WaterMainReadings.objects.filter(read_date__gte=one_month_ago, read_date__lte=current_day).count()
            monthly_all_consumptions_main.append(monthly_units)
            monthly_all_consumptions_clients.append(monthly_units_clients)
            monthss.append(this_month)
            #monthly_consumptions.append(this_month)
            current_day = current_day - datetime.timedelta(days=30)
            this_month = current_day.month
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
            current_month = current_month - datetime.timedelta(days=30)
        # print(credit_usage)
        context = {
            'messages_sent': monthly_all_consumptions[::-1],
            'monthlty_clients': monthly_all_consumptions_clientsw[::-1],
            'monthly_main': monthly_all_consumptions[::-1],
            #'weeks': weeks[::-1],
            'months': monthss[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'main_accounts': MainMeter.objects.filter().count(),
            'last_month_collections': WaterPaymentReceived.aggregate(Sum('amount')),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/main_meter.html', context)
    else:
        months = get_last_n_monther(10)
        weeks = get_last_n_weeks(20)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        this_month = current_day.month
        monthly_all_consumptions = []
        monthss=[]
        for month in months:
            #print(month)
            #ModelName.objects.aggregate(Sum('field_name'))
            consumed_units = WaterMainReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0
               
            messages = WaterMainReadings.objects.filter(read_date__gte=one_month_ago, read_date__lte=current_day).count()
            monthly_all_consumptions.append(consumed_units)
            monthss.append(this_month)
            #monthly_consumptions.append(this_month)
            current_day = current_day - datetime.timedelta(days=30)
            
            this_month = current_day.month
            last_month = this_month - 1
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
        this_day = datetime.datetime.now()
        today_month = this_day.month
        yesterday_month = today_month - 1
            
       
        context = {
            'messages_sent': monthly_all_consumptions[::-1],
            'months': monthss[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'courts': WaterCourt.objects.filter().count(),
            'readings': WaterMeterReadings.objects.filter().count(),
            'outbox': WaterMeterReadings.objects.filter().count(),
            'main_accounts': WaterNetwork.objects.filter().count(),
            'this_month' : this_month,
            'this_month_collections': WaterPaymentReceived.objects.filter(pay_date__month=today_month).aggregate(total=Sum('amount'))['total'] or 0,
            'last_month_collections': WaterPaymentReceived.objects.filter(pay_date__month=yesterday_month).aggregate(total=Sum('amount'))['total'] or 0,
            'tobe_collected': WaterClientAll.objects.filter().aggregate(total=Sum('amount_due'))['total'] or 0,
            'unallocated_payments': MiwamaMpesa.objects.filter(processed=2).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/main_meter.html', context)






@login_required()
@is_user_customer
def water_apps(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is None:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        #weeks = get_last_n_weeks(12)
        months = get_last_n_months(10)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        this_month = current_day.month
        monthly_all_consumptions = []
        
        for month in months:
            # print(week)
            monthly_units = WaterMeterReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0
               
            messages = WaterMeterReadings.objects.filter(read_date__gte=one_month_ago, read_date__lte=current_day).count()
            monthly_all_consumptions.append(monthly_units)
            monthss.append(this_month)
            #monthly_consumptions.append(this_month)
            current_day = current_day - datetime.timedelta(days=30)
            this_month = current_day.month
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
            current_month = current_month - datetime.timedelta(days=30)
        # print(credit_usage)
        context = {
            'messages_sent': monthly_all_consumptions[::-1],
            #'weeks': weeks[::-1],
            'months': monthss[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/water_apps.html', context)
    else:
        months = get_last_n_monther(10)
        weeks = get_last_n_weeks(20)
        one_month_ago = datetime.datetime.today() - datetime.timedelta(days=30)
        current_day = datetime.datetime.today()
        this_month = current_day.month
        monthly_all_consumptions = []
        monthss=[]
        for month in months:
            #print(month)
            #ModelName.objects.aggregate(Sum('field_name'))
            consumed_units = WaterMeterReadings.objects.filter(read_date__month=this_month).aggregate(total=Sum('units_consumed'))['total'] or 0
               
            messages = WaterMeterReadings.objects.filter(read_date__gte=one_month_ago, read_date__lte=current_day).count()
            monthly_all_consumptions.append(consumed_units)
            monthss.append(this_month)
            #monthly_consumptions.append(this_month)
            current_day = current_day - datetime.timedelta(days=30)
            this_month = current_day.month
            one_month_ago = one_month_ago - datetime.timedelta(days=30)
            
       
        context = {
            'messages_sent': monthly_all_consumptions[::-1],
            'months': monthss[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'courts': WaterCourt.objects.filter().count(),
            'readings': WaterMeterReadings.objects.filter().count(),
            'outbox': WaterMeterReadings.objects.filter().count(),
            'unallocated_payments': MiwamaMpesa.objects.filter(processed=3).count(),
            'unallocated_amount': int(MiwamaMpesa.objects.filter(processed=3).aggregate(total=Sum('amount'))['total'] or 0),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/water_apps.html', context)
@login_required()
@is_user_customer


@login_required()
@is_user_customer
def water_apps_orig(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is None:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        weeks = get_last_n_weeks(12)
        months = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/water_apps.html', context)
    else:
        weeks = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'water_clients': WaterClientAll.objects.filter().count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'courts': WaterCourt.objects.filter().count(),
            'readings': WaterMeterReadings.objects.filter().count(),
            'outbox': WaterMeterReadings.objects.filter().count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/water_apps.html', context)
@login_required()
@is_user_customer
def apps(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is None:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        weeks = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/apps.html', context)
    else:
        weeks = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/apps.html', context)


@login_required()
@is_user_customer
def monthly_messages_sent(request):
    return


@login_required()
@is_user_customer
def services(request):
    return render(request, 'sms/services.html')


@login_required()
@is_user_customer
def contact(request):
    return render(request, 'sms/contact.html')


# @login_required()
class SendSmsView(LoginRequiredMixin, CreateView):
    model = Outgoing
    fields = ['phone_numbers', 'text_message']

    def form_valid(self, form):
        m = form.cleaned_data['phone_numbers'].splitlines()

        # for n in form.cleaned_data['phone_numbers'].splitlines():
        for n in m:
            form.instance.phone_numbers=n
            form.instance.text_message=form.cleaned_data['text_message']
            form.instance.user = self.request.user
            form.instance.access_code=self.request.user.profile.access_code
            form.instance.service_id=self.request.user.profile.service_id
        return super().form_valid(form)


# @login_required()
class SmsReportView(ListView):
    model = Outgoing

from functools import wraps
from time import time



@timed
@login_required()
@is_user_customer
def send(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        credit = customer.credit
        if request.method == 'POST':
            data = request.session['simple_messages']
            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                return redirect('sms:customer_reports', track_code)
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'sms/simple_message.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_message.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        credit = customer.credit
        if request.method == 'POST':
            data = request.session['simple_messages']
            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                return redirect('sms:customer_reports', track_code)
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'sms/simple_message.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_message.html', context)


@task
def simple_sms_store(customer_id, total_message_cost, data, track_code):
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    bulk_mgr = BulkCreateManager(chunk_size=100)
    for a, b in data.items():
        s = ''.join(a.split())
        p = f"{254}{s[-9:]}"

        # if str.isdigit(customer.access_code):
        #     bulk_mgr.add(Outgoing(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=p,
        #         text_message=b,
        #         track_code=track_code
        #     ))
        # else:
        bulk_mgr.add(OutgoingNew(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=p,
            text_message=b,
            track_code=track_code,
            sent_time=timezone.now()
        ))
    bulk_mgr.done()
    send_usage(total_message_cost)
    return 'insertion complete'


@login_required()
@is_user_customer
def simple_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()
            data = {}
            # print(phone_numbers)
            new_phone_numbers = []
            for phone_number in phone_numbers:
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    # pprint(p)
                    data[p] = message
                # pprint(data)
                request.session['simple_messages'] = data
                return redirect("sms:simple_sms_preview")
            else:
                messages.error(request, 'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_sms.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()
            data = {}
            # print(phone_numbers)
            new_phone_numbers = []
            for phone_number in phone_numbers:
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    # pprint(p)
                    data[p] = message
                # pprint(data)
                request.session['simple_messages'] = data
                return redirect("sms:simple_sms_preview")
            else:
                messages.error(request,
                               'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_sms.html', context)


@login_required()
@is_user_customer
def simple_sms_preview(request):
    # print(request.session['simple_messages'])
    context = {
        'data': request.session['simple_messages']
    }
    return render(request, 'sms/simple_sms_preview.html', context)


@login_required()
@is_user_customer
def import_csv_2(request):
    if request.method == 'POST' and request.FILES['myfile']:
        file = request.FILES['myfile']
        context = get_excel_content(file)
        return render(request, 'sms/sms.html', context)
    return render(request, 'sms/sms.html')
def import_readings_csv(request):
    if request.method == 'POST' and request.FILES['myfile']:
        file = request.FILES['myfile']
        context = get_excel_content(file)
        return render(request, 'sms/readings.html', context)
    return render(request, 'sms/readings.html')

@login_required()
@is_user_customer
def messages_dashboard(request):
    return render(request, 'sms/dashboard.html')


@login_required()
@is_user_customer
def merge_sms_2(request):
    data = []
    if request.method == 'POST':
        message = request.POST['Message']
        phone_number_field = request.POST['NumberField']
        file = request.POST['file_path']

        file_path = file.split('/', 1)[1]
        workbook = load_workbook('media/%s' % file_path)
        sheet_names = workbook.sheetnames

        sheet = sheet_names[0]
        worksheet = workbook.get_sheet_by_name(sheet)

        parameters = get_message_parameters(message=message)
        parameter_cells = get_parameter_column(parameters=parameters, worksheet=worksheet)
        phone_number_column = get_phone_number_column(phone_number_field=phone_number_field, worksheet=worksheet)

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        for i in range(2, max_row + 1):
            person_message = {}
            new_message = message
            sms = ''
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)

                for a, b in parameter_cells.items():
                    if j == b:
                        new_message = new_message.replace('[%s]' % a, str(cell_obj.value))
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if j == phone_number_column:
                    phone_number = cell_obj.value
                    sms = Merged(phone_number, new_message)
                    person_message['phone_number'] = phone_number
                    person_message['message'] = new_message
            data.append(sms)
    data_dict = {}
    for d in data:
        data_dict[d.phone_number] = d.message
    request.session['data'] = data_dict
    return render(request, 'sms/sample_merged_sms.html', {'data': data})


@login_required()
@is_user_customer
def confirm(request):
    # pprint(request.session['data'])
    if request.method == 'POST':
        tracking_code = random.randint(1, 1000000)
        while OutgoingNew.objects.filter(track_code=tracking_code).count() > 0:
            tracking_code = random.randint(1, 1000000)
        phone_numbers = request.POST.getlist('phone_numbers[]')
        c_messages = request.session['data']
        # pprint(c_messages)
        customer = Customer.objects.filter(user_ptr_id=request.user).first()
        if customer is not None:
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"

                    # if str.isdigit(customer.access_code):
                    #     bulk_mgr.add(Outgoing(
                    #         customer=customer,
                    #         service_id=customer.service_id,
                    #         access_code=customer.access_code,
                    #         phone_number=p,
                    #         text_message=message,
                    #         track_code=tracking_code
                    #     ))
                    # else:
                    bulk_mgr.add(OutgoingNew(
                        customer=customer,
                        service_id=customer.service_id,
                        access_code=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        sent_time=timezone.now()
                    ))

                    # bulk_mgr.add(Outgoing(phone_number=p,
                    #                       text_message=message,
                    #                       service_id=customer.service_id,
                    #                       access_code=customer.access_code,
                    #                       customer_id=customer.id,
                    #                       track_code=tracking_code))
                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                send_usage(actual_messages_cost)
                return redirect('sms:customer_reports', tracking_code)
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'sms/sample_merged_sms.html', {'data': data})
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user).first().owner
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"
                    # bulk_mgr.add(Outgoing(phone_number=p,
                    #                       text_message=message,
                    #                       service_id=customer.service_id,
                    #                       access_code=customer.access_code,
                    #                       customer_id=customer.id,
                    #                       track_code=tracking_code))
                    #
                    # if str.isdigit(customer.access_code):
                    #     bulk_mgr.add(Outgoing(
                    #         customer=customer,
                    #         service_id=customer.service_id,
                    #         access_code=customer.access_code,
                    #         phone_number=p,
                    #         text_message=message,
                    #         track_code=tracking_code
                    #     ))
                    # else:
                    bulk_mgr.add(OutgoingNew(
                        customer=customer,
                        service_id=customer.service_id,
                        access_code=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        sent_time=timezone.now()
                    ))
                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                send_usage(actual_messages_cost)
                return redirect('sms:customer_reports', tracking_code)
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'sms/sample_merged_sms.html', {'data': data})
    return render(request, 'sms/result.html')


def register(request):
    if request.method == "POST":
        sdp = SDP()
        customer_code = random.randint(10000, 99999)
        while Customer.objects.filter(customer_code=customer_code).count() > 0:
            customer_code = random.randint(10000, 99999)
        form = CustomerRegisrationForm(request.POST)
        if form.is_valid():
            if Customer.objects.filter(username=request.POST['email']).count() < 1:
                customer = form.save(commit=False)
                customer.first_name = request.POST['username']
                customer.username = request.POST['email']
                customer.email = request.POST['email']
                customer.is_active = False
                customer.customer_code = customer_code
                customer.access_code = 'ROBERMS_LTD'
                customer.sender_name = 'ROBERMS_LTD'
                customer.save()
                phone_number = f"{254}{customer.phone_number.replace(' ', '')[-9:]}"
                OutgoingNew.objects.create(phone_number=phone_number,
                                      text_message="Account created successfully!\n" +
                                                   f"Verify the account with this code {customer.customer_code} now.\n" +
                                                   f"If you are verifying later, use the link below to activate \n" +
                                                   f"https://roberms.co.ke/verify/account",
                                      service_id=6015152000175328,
                                      access_code='ROBERMS_LTD',
                                      customer_id=customer.id,
                                      track_code=customer_code,
                                           sent_time=timezone.now())
                # response = sdp.send_sms_customized(service_id='6015152000175328', recipients=[phone_number],
                #                                	message=f'Verfication Code {customer.customer_code}', sender_code='711037')
                messages.success(request, 'Your account has been created! We sent you an text message containing a verification code')
                return redirect('sms:verify_account')
            else:
                messages.error(request, 'That Email Is Already Registered To Our System')
                return render(request, 'registration/register.html')
        else:
            # print('me')
            # print(form.errors)
            return render(request, 'registration/register.html', {'form': form})
    else:
        form = CustomerRegisrationForm()
        return render(request, 'registration/register.html', {'form': form})


def verify_account(request):
    if request.method == 'POST':
        if Customer.objects.filter(customer_code=request.POST['customer_code'], is_active=False).count() > 0:
            customer = Customer.objects.filter(customer_code=request.POST['customer_code']).first()
            customer.is_active = True
            customer.save()
            messages.success(request, f'Welcome {customer.first_name} your account is now active')
            return redirect('sms:login')
        else:
            messages.error(request, 'That code verification is not valid, please try again ')
            return render(request, 'registration/verify_account.html')
    return render(request, 'registration/verify_account.html')


@login_required()
@is_user_customer
def profile(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    return render(request, 'sms/profile.html', {'customer': customer})


@login_required()
@is_user_customer
def edit_profile(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if request.method == 'POST':
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        business_name = request.POST['business_name']
        customer.email = email
        customer.first_name = first_name
        customer.last_name = last_name
        customer.business_name = business_name
        customer.save()
        messages.success(request, 'Profile Updated Successfully')
        return redirect('sms:apps')
    return render(request, 'sms/edit_profile.html', {'customer': customer})


def customer_credit(request):

    return render(request, 'sms/credit.html')


@login_required()
@is_user_customer
def reports(request, tracking_code):
    sent_messages = OutgoingDone.objects.filter(delivery_status__icontains='DeliveredTo', track_code=tracking_code).count()
    messages_not_sent = OutgoingDone.objects.filter(Q(delivery_status__icontains="InvalidMsisdn")|
                                                    Q(delivery_status__icontains="DeliveryImpossible")|
                                                    Q(delivery_status__icontains="AbsentSubscriber")|
                                                    Q(delivery_status__icontains="SenderName Blacklisted"),
                                                    track_code=tracking_code).count()
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': OutgoingDone.objects.filter(customer_id=customer.id, track_code=tracking_code, delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)
    else:
        sub_acc = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first()
        customer = sub_acc.owner
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': OutgoingDone.objects.filter(customer_id=customer.id, track_code=tracking_code,
                                                           delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)

@login_required()
@is_user_customer
def water_sms_reports(request, tracking_code):
    sent_messages = WaterOutbox.objects.filter(delivery_status__icontains='DeliveredTo', track_code=tracking_code).count()
    messages_not_sent = WaterOutbox.objects.filter(Q(delivery_status__icontains="InvalidMsisdn")|
                                                    Q(delivery_status__icontains="DeliveryImpossible")|
                                                    Q(delivery_status__icontains="AbsentSubscriber")|
                                                    Q(delivery_status__icontains="SenderName Blacklisted"),
                                                    track_code=tracking_code).count()
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': WaterOutbox.objects.filter(track_code=tracking_code, delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)
    else:
        sub_acc = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first()
        customer = sub_acc.owner
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': OutgoingDone.objects.filter(track_code=tracking_code,
                                                           delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)


@login_required()
@is_user_customer
def customer_contacts(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)

def water_courts(request):
    customer = Customer.objects.filter(id=request.user.id).first()
    if customer is not None:
        networks = WaterNetwork.objects.all()

        context = {

            'networks': networks
        }
        return render(request, 'sms/courts.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)
#0702455926
def water_clienter(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/water_clients.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/water_clients.html', context)
def water_clients(request):
    clients = WaterClientAll.objects.all().order_by('-id').values()
    context = {
        'clients': clients
    }
    return render(request, 'sms/water_clients.html', context)
def water_clients_court(request,court_id):
    court = WaterCourt.objects.filter(id=court_id).first()
    court_name = court.name
    clients = WaterClientAll.objects.filter(court=court_name).order_by('-id').values()
    context = {
        'clients': clients
    }
    return render(request, 'sms/water_clients.html', context)
def water_reports(request):
    clients = WaterClientAll.objects.all()
    context = {
        'clients': clients
    }
    return render(request, 'sms/water_reports.html', context)
def water_payments(request):
    payments = WaterPaymentReceived.objects.all().order_by('-pay_date')[:600]
    context = {
        'payments': payments
    }
    return render(request, 'sms/water_payments.html', context)
def water_payments_unallocated(request):
    payments = MiwamaMpesa.objects.filter(processed=3).order_by('-id').values()
    context = {
        'payments': payments
    }
    return render(request, 'sms/water_payments_unallocated.html', context)
def water_payments_clients(request,client_id):
    payments = WaterPaymentReceived.objects.filter(account_number=client_id)
    context = {
        'payments': payments
    }
    return render(request, 'sms/water_payments.html', context)

def client_invoicers(request, client_id):
    client = Client.objects.get(id=client_id)
    invoices = WaterClientAll.objects.filter(client=client)

    context = {
        'client': client,
        'invoices': invoices
    }
    return render(request, 'invoices/client_invoices.html', context)

def edit_water_client(request, client_id):
    client = WaterClientAll.objects.get(id=client_id)
    amount_due1 = client.amount_due
    readings1 = client.last_meter_reading
    #WaterMeterReadings.objects.all().delete()
    #WaterPaymentReceived.objects.all().delete()
    #WaterPaymentReceivedManual.objects.all().delete()
    #WaterOutbox.objects.all().delete()
    #WaterMeterReplacement.objects.all().delete()
    #WaterPaymentReallocate.objects.all().delete()
    #WaterNetwork.objects.all().delete()
    #WaterCourt.objects.all().delete()
    #WaterClientAll.objects.all().delete()
    #MiwamaMpesa.objects.all().delete()

    if request.method == 'POST':
        client.names = request.POST['names']
        #client.msisdn = request.POST['msisdn']
        phones = request.POST['msisdn']
        phones2 = request.POST['msisdn2']
        new_balance = int(float(request.POST['amount_due']))
        phone_number = f"{0}{phones.replace(' ', '')[-9:]}"
        phone_number2 = f"{0}{phones2.replace(' ', '')[-9:]}"
        client.msisdn = phone_number
        client.msisdn2 = phone_number2
        client.court = request.POST['court']
        client.email_address = request.POST['email_address']
        client.last_meter_reading = int(float(request.POST['last_meter_reading']))
        client.customer_rate = int(float(request.POST['customer_rate']))
        client.standing_charge = int(float(request.POST['standing_charge']))
        client.amount_due = int(float(request.POST['amount_due']))
        client.save()
        if amount_due1 != client.amount_due:
            if amount_due1>client.amount_due:
                credit=amount_due1-client.amount_due
                debit=0
            else:
                           
                debit=client.amount_due-amount_due1
                credit=0
            WaterStatement.objects.create(
            account_number=client,
            narration='System update via client edit',
            debit=debit,
            curr_reading=request.POST['last_meter_reading'],
            prev_reading=readings1,
            credits = credit,
            balance = request.POST['amount_due']


        )

        #WaterNetwork.delete(self)
        #messages.success(request, request.POST['email_address'])
        messages.success(request,new_balance)
        return redirect('sms:water_clients')
    context = {
        'client': client
    }
    return render(request, 'sms/edit_water_client.html', context)

def edit_water_network(request, network_id):
    client = WaterNetwork.objects.get(id=network_id)
    if request.method == 'POST':
        client.standing_charge = request.POST['standing_charge']
        client.rate = request.POST['rate']
        
        client.save()
        #WaterNetwork.delete(self)

       

        messages.success(request, request.POST['rate'])
        return redirect('sms:water_courts')
    context = {
        'client': client
    }
    return render(request, 'sms/edit_network.html', context)

def edit_sys_config(request, client_id):
    client = WaterSysConf.objects.get(id=client_id)
    if request.method == 'POST':
        client.standing_charge = request.POST['standing_charge']
        client.rate = request.POST['rate']
        client.water_levy = request.POST['water_levy']
        client.comment = request.POST['comment']
        client.save()
        #WaterNetwork.delete(self)

        WaterSysConfHist.objects.create(
            standing_charge=request.POST['standing_charge'],
            rate=request.POST['rate'],
            water_levy=request.POST['water_levy'],
            comment = request.POST['comment']


        )

        messages.success(request, request.POST['rate'])
        return redirect('sms:water_system_config')
    context = {
        'client': client
    }
    return render(request, 'sms/edit_sys_config.html', context)

def meter_readings(request):
    #meter_readings = WaterMeterReadings.objects.all()
    meter_readings = WaterMeterReadings.objects.all().order_by('-read_date')[:600]
    #meter_readingss = WaterMeterReadings.objects.all().delete()

    context = {
        'meter_readings': meter_readings
    }
    return render(request, 'sms/water_meter_readings.html', context)

def main_readings(request):
    meter_readings = WaterMainReadings.objects.all()
    #meter_readingss = WaterMainReadings.objects.all().delete()

    context = {
        'meter_readings': meter_readings
    }
    return render(request, 'sms/water_main_readings.html', context)

def bills_sent(request):
    bills_sent = WaterBillSent.objects.all()
    #meter_readingss = WaterMeterReadings.objects.all().delete()

    context = {
        'bills_sent': bills_sent
    }
    return render(request, 'sms/bills_sent.html', context)

def meter_readings_clients(request,client_id):
    meter_readings = WaterMeterReadings.objects.filter(account_number=client_id)
    #meter_readingss = WaterMeterReadings.objects.all().delete()

    context = {
        'meter_readings': meter_readings
    }
    return render(request, 'sms/water_meter_readings.html', context)

def bills_sent_clients(request,client_id):
    bills_sent = WaterBillSent.objects.filter(account_number=client_id)
    #meter_readingss = WaterMeterReadings.objects.all().delete()

    context = {
        'bills_sent': bills_sent
    }
    return render(request, 'sms/bills_sent.html', context)

def import_fine_readings(request):
    if request.method == 'POST':
        #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        customer = Customer.objects.filter(id=1).first()
        

        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            #f_path = "templates"
            extension = file.name.rsplit('.', 1)[1]
            workbook = load_workbook(filename=f_path, read_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            for i in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=i, column=4).value != 0:

                    account_number = worksheet.cell(row=i, column=1).value
                    name = worksheet.cell(row=i, column=2).value
                    fines = (worksheet.cell(row=i, column=4).value)
                    acount_client = WaterClientAll.objects.filter(id=account_number).first()
                    #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
                    client_id=acount_client.id

                    
                    #readings = float(worksheet.cell(row=i, column=3).value)
                    #readings = str(worksheet.cell(row=i, column=2).value)


                    try:
                        WaterFines.objects.create(
                            names=name,
                            fine_amount=fines,
                            account_number=acount_client,
                            reading_type="excel upload"

                        )
                    except TypeError:
                        continue
                    except ValueError:
                        continue
                    print('saved')
            messages.success(request, filename)
            fines = WaterFines.objects.filter(processed=0)
           # meter_readings = WaterMeterReadings.objects.all().delete()
            # Book.objects.all().delete()
            context = {
                'fines': fines
            }
            return render(request, 'sms/import_fine_readings.html',context)

        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_meter_readings_task.delay(extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)


def disconnection_reminder(request, client_id):
    client = WaterClientAll.objects.get(id=client_id)
    client_name=client.names
    client_phone=client.msisdn
    account_client=str(client.id)
    amountd=client.amount_due
    client_balance=str(client.amount_due)
    dear = "Dear "
    final = ", You are listed for disconnection due to your outstanding water bill of Ksh."
    payment = ". clear your bill through our Paybill 499086, account "
    disconnection = " to avoid disconnection."
    client_message = dear  +  client_name + final + client_balance + payment + account_client  + disconnection
       
    if request.method == 'POST':
        client_name=client.names
        client_phone=client.msisdn
        client_balance=client.amount_due
        client_message=client_message
        #client_message = 'Dear' + client_name, 'This is a final reminder to clear your outstanding water bill of Ksh.' + client_balance + ' by end of today. Please note, disconnection of normal supply will be done if we do not hear from you.aqua nova management.'
       
        WaterOutbox.objects.create(
            dest_msisdn=client_phone,
            text_message=client_message,
            user_id=100,
            client=client_id
            


        )

        messages.success(request, "Message resent sucessfully")
        return redirect('sms:water_sent_sms')
    context = {
        'client': client,
        'reminder_message':client_message,
        'client_bal':amountd,
        'low_balance':'The client has an arrears of Less than 100 hence cannot be disconnected'
    }
    return render(request, 'sms/reminder_message.html', context)



def import_meter_readings(request):
    if request.method == 'POST':
        #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        customer = Customer.objects.filter(id=1).first()
        def meter_readings1():
            WaterMeterReadingsRaw.objects.all().delete()
        meter_readings1()

        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            #f_path = "templates"
            extension = file.name.rsplit('.', 1)[1]
            workbook = load_workbook(filename=f_path, read_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            for i in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=i, column=4).value != 0:

                    account_number = worksheet.cell(row=i, column=1).value
                    name = worksheet.cell(row=i, column=2).value
                    readings = (worksheet.cell(row=i, column=4).value)
                    #readings = float(worksheet.cell(row=i, column=3).value)
                    #readings = str(worksheet.cell(row=i, column=2).value)


                    try:
                        WaterMeterReadingsRaw.objects.create(
                            names=name,
                            readings=readings,
                            account_number=account_number,
                            reading_type="excel upload"

                        )
                    except TypeError:
                        continue
                    except ValueError:
                        continue
                    print('saved')
            messages.success(request, filename)
            meter_readings = WaterMeterReadingsRaw.objects.all()
           # meter_readings = WaterMeterReadings.objects.all().delete()
            # Book.objects.all().delete()
            context = {
                'meter_readings': meter_readings
            }
            return render(request, 'sms/water_import_meter_reading.html',context)

        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_meter_readings_task.delay(extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)


@task()
def store_meter_readings_task():
    WaterMeterReadings.objects.update_or_create(
        names="Benard Barongo",
        msisdn="724648426",
        account_number="2",
        previous_reading="12",
        readings=14,
        id_num="2343",
        reading_type="System Input",
        units_consumed=2

    )
    # CustomerTask.objects.filter(task_id=store_contact_task.id).update(
    #     status_complete=True
    # )
    return 'completed'


def create_client(request):
    if request.method == 'POST':
        customer_number = ''
        last_client = Client.objects.all().order_by('id').last()
        if not last_client:
            customer_number = 'TNT-100'
        else:
            cn = last_client.client_number
            cn_int = int(cn.split('TNT-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"TNT-{new_cn_int}"
            customer_number = new_cn
        Client.objects.create(
            company_name=request.POST['company_name'],
            phone_number=request.POST['phone_number'],
            client_number=customer_number,
            kra_pin=request.POST['kra_pin'],
            address=request.POST['address'],
            location=request.POST['location']
        )
        messages.success(request, 'Clienter Added Successfully')
        return redirect('Invoices:invoice_clients')
    return render(request, 'sms/create_client.html')



def add_fines(request):

    if request.method == 'POST':
        fines = int(request.POST['fines'])
        account_num = request.POST['client']


        q = WaterClientAll.objects.filter(id=account_num)
        sys_configs = WaterSysConf.objects.filter().first()
        standing_charge=sys_configs.standing_charge
        rate=sys_configs.rate
        waterclient = WaterClientAll.objects.get(id=account_num)
        names = q[0].names
        the_ids = q[0].id
        msisdn = q[0].msisdn
        id_num = q[0].id_num
        amount_due = q[0].amount_due
        amount_0 = q[0].amount_0
        amount_1 = q[0].amount_1
        amount_2 = q[0].amount_2

        
        if fines>0:
            



            WaterFines.objects.create(
                
                account_number=waterclient,
                fine_amount=fines,
                names=waterclient.names,
                reading_type='System input'
                

            )

            

            messages.success(request,  names)
            return redirect('sms:add_fines')
        else:
            print("Invalid fines")
            messages.success(request, "Invalid fines")
            clienter = WaterClientAll.objects.all()
            fines = WaterFines.objects.all()
            context = {
                'clients': clienter,
                'fines' : fines
            }
            return render(request, 'sms/add_fines.html', context)
    else:

        clienter = WaterClientAll.objects.all()
        fines = WaterFines.objects.all()
        context = {
            'clients': clienter,
            'fines' : fines
        }

        #return render(request, 'sms/add_meter_readings.html', context)
        return render(request, 'sms/add_fines.html', context)


def add_main_readings(request):

    if request.method == 'POST':
        readings = int(request.POST['readings'])
        meter_id = request.POST['customers']
        last_m_readings = 0

        main_meter = WaterNetwork.objects.get(id=meter_id)
        last_readings = WaterMainReadings.objects.filter().last()
        
        last_m_readings = main_meter.reading
        units_consumed=readings - last_m_readings

        
        if units_consumed>0:
            main_meter.reading = readings
            main_meter.read_date=datetime.datetime.now()
            main_meter.save()

            WaterMainReadings.objects.create(
            meter_num=meter_id,
            network=main_meter.name,
            readings=readings,
            units_consumed=units_consumed

        )
        else:
            print("Invalid readings")
            messages.success(request, "Invalid readings for the main meter Previous readings were " + str(last_m_readings))
            clienter = WaterClientAll.objects.all()
            context = {
                'meter': MainMeter.objects.all()
            }
            return render(request, 'sms/add_main_readings.html', context)


        context = {
            'meter': MainMeter.objects.all()
        }
        return render(request, 'sms/add_main_readings.html', context)
    else:
        context = {
            'meter': WaterNetwork.objects.all()
        }
        
        return render(request, 'sms/add_main_readings.html', context)

def add_meter_readings(request):

    if request.method == 'POST':
        readings = request.POST['readings']
        account_num = request.POST['client']


        q = WaterClientAll.objects.filter(id=account_num)
        sys_configs = WaterSysConf.objects.filter().first()
        standing_charge=sys_configs.standing_charge
        rate=sys_configs.rate
        waterclient = WaterClientAll.objects.get(id=account_num)
        names = q[0].names
        the_ids = q[0].id
        msisdn = q[0].msisdn
        id_num = q[0].id_num
        amount_due = q[0].amount_due
        amount_0 = q[0].amount_0
        amount_1 = q[0].amount_1
        amount_2 = q[0].amount_2

        last_meter_reading = q[0].last_meter_reading
        units_consumed=int(readings)-int(last_meter_reading)
        if units_consumed>0:
            amount_from_units=(units_consumed*rate)+standing_charge



            WaterMeterReadings.objects.create(
                names=names,
                msisdn=msisdn,
                account_number=waterclient,
                previous_reading=last_meter_reading,
                readings=readings,
                id_num=id_num,
                reading_type="System Input",
                units_consumed=units_consumed


            )

            waterclient.units_consumed = units_consumed
            waterclient.last_meter_reading = readings
            waterclient.amount_0=amount_from_units
            waterclient.amount_1 = amount_0
            waterclient.amount_2 = amount_1
            waterclient.amount_3 = amount_2

            waterclient.last_meter_reading_date = datetime.datetime.now()

            waterclient.save()

            messages.success(request,  names)
            return redirect('sms:meter_readings')
        else:
            print("Invalid readings")
            messages.success(request, "Invalid readings for " + names + ". Previous readings were " + str(last_meter_reading))
            clienter = WaterClientAll.objects.all()
            context = {
                'clients': clienter
            }
            return render(request, 'sms/add_meter.html', context)
    else:

        clienter = WaterClientAll.objects.all()
        context = {
            'clients': clienter
        }

        #return render(request, 'sms/add_meter_readings.html', context)
        return render(request, 'sms/add_meter.html', context)


def add_fine_upload(request):
    customer = WaterClientAll.objects.filter(amount_due__gt=0).order_by('id')
    time = datetime.datetime.now()
    file_namer = "fines"
    file_path = 'media/reports/'
    filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
    #filename = "meter_readings.xlsx" % ( time.year, time.month)
    full_path = f"{file_path}/{filename}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook = Workbook()
    summary_sheet = workbook.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Fines template'
    #cell = summary_sheet.cell(row=1, column=1)
    #cell.value = 'Meter Readings'
    #cell.alignment = Alignment(horizontal='center', vertical='center')
    #summary_sheet.merge_cells('A1:B1')
    summary_sheet.append(('A/C','NAMES', 'AMOUNT DUE','FINES'))

    number = 1
    for cust in customer:
        #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
        summary_sheet.append((cust.id, cust.names,cust.amount_due, 0))
        number += 1

    workbook.save(full_path)
    context = {
        'file_path': full_path
    }
    return render(request, 'sms/add_fines.html', context)
def add_meter_upload(request):
    customer = WaterClientAll.objects.all().order_by('id')
    time = datetime.datetime.now()
    file_namer = "meter_readings"
    file_path = 'media/reports/'
    filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
    #filename = "meter_readings.xlsx" % ( time.year, time.month)
    full_path = f"{file_path}/{filename}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook = Workbook()
    summary_sheet = workbook.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Meter Reading Template'
    #cell = summary_sheet.cell(row=1, column=1)
    #cell.value = 'Meter Readings'
    #cell.alignment = Alignment(horizontal='center', vertical='center')
    #summary_sheet.merge_cells('A1:B1')
    summary_sheet.append(('A/C','NAMES', 'PREV. READINGS','CURRENT READINGS','STATION'))

    number = 1
    for cust in customer:
        #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
        summary_sheet.append((cust.id, cust.names,cust.last_meter_reading, 0,cust.network))
        number += 1

    workbook.save(full_path)
    context = {
        'file_path': full_path
    }
    return render(request, 'sms/add_meter.html', context)


def meter_readings_report(request):
    #from datetime import datetime
    from datetime import timedelta
    from django.utils import formats
    if request.method == 'POST':
        report_value = request.POST['reports_value']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        current_datetime = datetime.datetime.now()
        #date_joined = datetime.now()  
        date_time = current_datetime.strftime("%Y-%m-%d")
        plus_one_day = datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1)
        formatted_datetime = formats.date_format(current_datetime, "SHORT_DATETIME_FORMAT")
        plus_one_months = datetime.datetime.strptime(date_time, "%Y-%m-%d") + datetime.timedelta(days=-31)
        #plus_one_months = formatted_datetime + datetime.timedelta(days=31)
        #end_date = datetime(end_date) + timedelta(days=1)
        m = request.POST['reports_value']
        #start_date = start_date.strptime(start_date, "YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ]")
        #end_date = end_date.strptime(end_date, "YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ]")
        if report_value=='1':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterClientAll.objects.filter(created_at__range=[start_date, plus_one_day]).order_by('id')
            report_net=2
            time = datetime.datetime.now()
            file_namer = "water_clients_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Water Clients'
            #summary_sheet.title = m
            #cell = summary_sheet.cell(row=1, column=1)
            #cell.value = 'Meter Readings'
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            wb = Workbook()  
            sheet = wb.active  
            sheet.merge_cells('A1:B2')  
  
            cell = sheet.cell(row=1, column=1)  
            cell.value = 'TOM WATER CLIENTS' 
            #summary_sheet.merge_cells('A2:D2')
            
            wb=Workbook()
            ws=wb.active
            a1=ws['A2']
            ft=Font(color="FA8072")
            a1.font=ft


















            #summary_sheet.append(('T N T WATER CLIENTS'))
            summary_sheet.append(('TOM WATER CLIENTS','NAMES'))
            big_red_text = Font(color="00FF0000", size=20)
            sheet["B1"].font = big_red_text
            sheet["B2"].font = big_red_text
            summary_sheet.append(('A/C','NAMES', 'PHONE NUMBER', 'ID NUMBER', 'STATION','REG DATE', 'READINGS','AMOUNT DUE'))
            from openpyxl.styles import PatternFill, GradientFill
            ws.sheet_view.showGridLines = False

            for c in ws['B5:C5'][0]:
                c.fill = PatternFill('solid', fgColor = 'F2F2F2')

            for c in ws['B7:C7'][0]:
                c.fill = PatternFill('gray0625')

            number = 1
            for cust in customer:
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                summary_sheet.append((cust.id, cust.names, cust.msisdn,cust.id_num,cust.network,cust.created_at ,cust.last_meter_reading,cust.amount_due))
                number += 1
            


                double = Side(border_style="double", color="4617F1")
                thin = Side(border_style="thin", color="4617F1")
                regular = Side(border_style="medium", color="000000")

                ## For the title cells B2 to F2
                for c in ws['B4:H4'][0]:
                    c.border = Border(bottom=double, top=thin)



                no_left_side = Border(top = regular,bottom=regular,right=regular)
                no_right_side = Border(top = regular,bottom=regular, left=regular)
                box = Border(top = regular,bottom=regular, left=regular,right=regular)

                ## For the "table-like" cells
                for c in ws['B8:B11']+ws['B15:B19']:
                    c[0].border = no_left_side
                    
                for c in ws['C8:C11']+ws['C15:C19']:
                    c[0].border = no_right_side

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)
        if report_value=='4':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterPaymentReceived.objects.filter(pay_date__range=[start_date, plus_one_day]).order_by('-id')
            report_net=2
            time = datetime.datetime.now()
            file_namer = "payment_received_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Received Payments'
            #summary_sheet.title = m
            #cell = summary_sheet.cell(row=1, column=1)
            #cell.value = 'Meter Readings'
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            #summary_sheet.merge_cells('A1:B1')
            summary_sheet.append(('PAYMENT DATE','A/C NUMBER', 'ACCOUNT NAME', 'AMOUNT', 'PAID BY','CONFIRMATION CODE', 'OUTSTANDING BALANCE','STATION'))

            number = 1
            for cust in customer:
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                summary_sheet.append((cust.pay_date, cust.account_number, cust.account_name,cust.amount,cust.received_from,cust.confirmation_code ,cust.balance_carried_forward,cust.client.network))
                number += 1

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)
        if report_value=='5':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterClientAll.objects.filter(amount_due__gt=0).order_by('id')
            #customer = WaterClientAll.objects.filter(id=1).order_by('id')
            report_net=2
            time = datetime.datetime.now()
            file_namer = "arrears_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = 'ARREARS REPORT'
            #summary_sheet.title = m

            cell = summary_sheet.cell(row=2, column=5)


            cell.value = 'CUMMULATIVE ARREARS'
            cell.alignment = Alignment(horizontal='center', vertical='center')



            summary_sheet.merge_cells('E2:h2')
            summary_sheet.append(('A/C','NAMES', 'PHONE NUMBER' ,'STATION','CURRENT ARREARS', '1 MONTH OLD', '2 MONTHS OLD','3 MONTHS', 'OVER 3 MONTHS','TOTAL ARREARS'))

            number = 1
            for cust in customer:

                cust_id=cust.id
                #last_readings = WaterMeterReadings.objects.filter(account_number=cust_id).order_by('id').last()
                #last_readings = WaterMeterReadings.objects.get(id=143)
                payable0=cust.amount_0
                payable1=cust.amount_1
                payable2=cust.amount_2
                payable3=cust.amount_3
                last_arrears=cust.amount_0
                total_arrears=cust.amount_due

                m2_arrears=0
                m3_arrears=0
                m4_arrears=0
                m0_arrears=0
                m1_arrears=0
                    




                if last_arrears==0:
                    last_arrears=cust.amount_due


                two_months = 1000
                three_months=0
                current_arrears=0
                month_2_arrears=0
                month_3_arrears=0
                month_4_arrears=0


                if total_arrears<cust.amount_0:
                    m1_arrears=total_arrears
                    m2_arrears=0
                    m3_arrears=0
                    m4_arrears=0
                else:
                    #total_arrears>cust.amount_0
                    m1_arrears=total_arrears-cust.amount_0
                    if m1_arrears<0:
                        #m1_arrear=cust.amount_0-total_arrears
                        m1_arrear=0
                        m2_arrears=0
                        m3_arrears=0
                        m4_arrears=0
                    else: 
                        m2_arrears=m1_arrears-cust.amount_1
                        if m2_arrears<0:
                            m2_arrears=0
                            m3_arrears=0
                            m4_arrears=0
                        else:
                            m3_arrears-m2_arrears-cust.amount_2
                            if m3_arrears<0:
                                m3_arrears=0
                                m4_arrears=0
                            else:
                                m4_arrears=m3_arrears-cust.amount_3
                                if m4_arrears<0:
                                    m4_arrears=0
                        



                    if cust.amount_1==0:
                        m1_arrears=0
                        m2_arrears=0
                        m3_arrears=0
                        m4_arrears=0
                    if cust.amount_2==0:                        
                        m2_arrears=0
                        m3_arrears=0
                        m4_arrears=0
                    if cust.amount_3==0:                        
                        m3_arrears=0
                        m4_arrears=0





                if total_arrears>cust.amount_0:
                    month_1_arrears=total_arrears-cust.amount_0
                else:
                    month_1_arrears=total_arrears
                if cust.amount_0<1:
                    month_1_arrears=0

                if month_1_arrears>cust.amount_1:
                    month_2_arrears=month_1_arrears-cust.amount_1
                else:
                    month_2_arrears=0
                if month_2_arrears>cust.amount_2:
                    month_3_arrears=month_2_arrears-cust.amount_2
                else:
                    month_3_arrears=0
                if month_3_arrears>cust.amount_3:
                    month_4_arrears=month_3_arrears-cust.amount_3
                else:
                    month_4_arrears=0






                if last_arrears>=current_arrears:

                    current_arrears = last_arrears
                    
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                #summary_sheet.append((cust.id, cust.names, cust.msisdn,last_arrears,month_1_arrears,month_2_arrears,month_3_arrears ,month_4_arrears,cust.amount_due))
                summary_sheet.append((cust.id, cust.names, cust.msisdn,cust.network,cust.amount_due,m1_arrears,m2_arrears,m3_arrears,m4_arrears ,cust.amount_due))
                number += 1

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)
        if report_value=='6':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterClientAll.objects.filter(last_meter_reading_date__lte=plus_one_months).order_by('last_meter_reading_date')
            report_net=2
            time = datetime.datetime.now()
            file_namer = "overdue_meter_readings_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = "Overdue readings"
            #summary_sheet.title = m
            #cell = summary_sheet.cell(row=1, column=1)
            #cell.value = 'Meter Readings'
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            #summary_sheet.merge_cells('A1:B1')
            summary_sheet.append(('A/C','NAMES', 'PHONE NUMBER', 'ID NUMBER', 'STATION','LAST READING DATE', 'READINGS','AMOUNT DUE'))

            number = 1
            for cust in customer:
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                summary_sheet.append((cust.id, cust.names, cust.msisdn,cust.id_num,cust.network,cust.last_meter_reading_date,cust.last_meter_reading,cust.amount_due))
                number += 1

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)
            
        if report_value=='2':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterMeterReadings.objects.filter(read_date__range=[start_date, plus_one_day]).order_by('id')
            report_net=2
            time = datetime.datetime.now()
            file_namer = "meter_readings_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Meter Reading Report2'
            #summary_sheet.title = m
            #cell = summary_sheet.cell(row=1, column=1)
            #cell.value = 'Meter Readings'
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            #summary_sheet.merge_cells('A1:B1')
            summary_sheet.append(('A/C','NAMES','PHONE NUMBER', 'PR','CR', 'UNITS','BILL','CREDIT','ARREARS','PAYABLE', 'READING DATE', 'STATION'))

            number = 1
            for cust in customer:
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                summary_sheet.append((cust.account_number.id, cust.names,cust.msisdn, cust.previous_reading,cust.readings,cust.units_consumed,cust.amount_from_units,cust.credit,cust.arrears,cust.payable,cust.read_date,cust.account_number.network ))
                number += 1

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)
        if report_value=='3':


            #customer = WaterMeterReadings.objects.all().order_by('id')
            customer = WaterOutbox.objects.filter(out_date__range=[start_date, plus_one_day]).order_by('-id')

            time = datetime.datetime.now()
            file_namer = "Bulk_sms_logs_report"
            file_path = 'media/reports/'
            filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
            #filename = "meter_readings.xlsx" % ( time.year, time.month)
            full_path = f"{file_path}/{filename}"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Bulk sms logs'
            #summary_sheet.title = m
            #cell = summary_sheet.cell(row=1, column=1)
            #cell.value = 'Meter Readings'
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            #summary_sheet.merge_cells('A1:B1')
            summary_sheet.append(('OUT DATE','PHONE NUMBER', 'MESSAGE', 'STATUS'))

            number = 1
            for cust in customer:
                #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
                summary_sheet.append((cust.out_date, cust.dest_msisdn, cust.text_message,cust.delivery_status))
                number += 1

            workbook.save(full_path)
            context = {
                'file_path': full_path,
                'request_report':report_value
            }
            return render(request, 'sms/water_reports.html', context)






            #return redirect('sms:meter_readings_report')

        #return render(request, 'sms/meter_readings_report.html', context)
    return render(request, 'sms/water_reports.html')
def meter_readings_report_orig(request):
    #from datetime import datetime
    from datetime import timedelta
    if request.method == 'POST':
        report_value = request.POST['reports_value']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        plus_one_day = datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1)
        #end_date = datetime(end_date) + timedelta(days=1)
        m = request.POST['reports_value']
        #start_date = start_date.strptime(start_date, "YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ]")
        #end_date = end_date.strptime(end_date, "YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ]")

        #customer = WaterMeterReadings.objects.all().order_by('id')
        customer = WaterMeterReadings.objects.filter(read_date__range=[start_date, plus_one_day]).order_by('id')
        report_net=2
        time = datetime.datetime.now()
        file_namer = "meter_readings_report"
        file_path = 'media/reports/'
        filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
        #filename = "meter_readings.xlsx" % ( time.year, time.month)
        full_path = f"{file_path}/{filename}"
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        workbook = Workbook()
        summary_sheet = workbook.get_sheet_by_name('Sheet')
        summary_sheet.title = 'Meter Reading Report'
        #summary_sheet.title = m
        #cell = summary_sheet.cell(row=1, column=1)
        #cell.value = 'Meter Readings'
        #cell.alignment = Alignment(horizontal='center', vertical='center')
        #summary_sheet.merge_cells('A1:B1')
        summary_sheet.append(('A/C','NAMES', 'METER READINGS', 'UNITS COSUMED', 'READING DATE'))

        number = 1
        for cust in customer:
            #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
            summary_sheet.append((cust.id, cust.names, cust.readings,cust.units_consumed,cust.read_date ))
            number += 1

        workbook.save(full_path)
        context = {
            'file_path': full_path,
            'request_report':report_value
        }
        return render(request, 'sms/water_reports.html', context)
        #return render(request, 'sms/meter_readings_report.html', context)
    return render(request, 'sms/water_reports.html')

def meter_readings_report2(request,report_value):
    customer = WaterMeterReadings.objects.all().order_by('id')
    time = datetime.datetime.now()
    file_namer = "meter_readings_report"
    file_path = 'media/reports/'
    filename = "%s_%d_%d.xlsx" % (file_namer, time.microsecond, time.second)
    #filename = "meter_readings.xlsx" % ( time.year, time.month)
    full_path = f"{file_path}/{filename}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook = Workbook()
    summary_sheet = workbook.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Meter Reading Report'
    #cell = summary_sheet.cell(row=1, column=1)
    #cell.value = 'Meter Readings'
    #cell.alignment = Alignment(horizontal='center', vertical='center')
    #summary_sheet.merge_cells('A1:B1')
    summary_sheet.append(('A/C','NAMES', 'METER READINGS', 'UNITS COSUMED', 'READING DATE'))

    number = 1
    for cust in customer:
        #summary_sheet.append((cust.id, cust.names, cust.last_meter_reading))
        summary_sheet.append((cust.id, cust.names, cust.readings,cust.units_consumed,cust.read_date ))
        number += 1

    workbook.save(full_path)
    context = {
        'file_path': full_path
    }
    return render(request, 'sms/water_reports.html', context)
    #return render(request, 'sms/meter_readings_report.html', context)

def meter_reading_upload_status(request, task_id):
    context = {
        'task_id': task_id
    }
    return render(request, 'sms/meter_reading_upload_status.html', context)
#return redirect('sms:contacts_upload_status', s.id)



def create_water_client(request):
    if request.method == 'POST':
        customer_number = ''
        last_client = WaterClientAll.objects.all().order_by('id').last()
        #last_client = '1'
        if not last_client:
            customer_number = 'TNT-100'
            cn = 'TNT-100'
            court = int(request.POST['court'])

            q = WaterCourt.objects.filter(id=court)

            court_name= q[0].name
            network_id = q[0].network_id
            s = WaterNetwork.objects.filter(id=network_id)
            network = s[0].name





            #cn = last_client.client_number
            cn_int = int(cn.split('TNT-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"TNT-{new_cn_int}"
            customer_number = new_cn
            phones = request.POST['msisdn']
            phones2 = request.POST['msisdn2']
            phone_number = f"{0}{phones.replace(' ', '')[-9:]}"
            phone_number2 = f"{0}{phones2.replace(' ', '')[-9:]}"
            WaterClientAll.objects.update_or_create(
                names=request.POST['names'],
                msisdn=phone_number,
                msisdn2=phone_number2,
                client_number=customer_number,
                id_num=request.POST['id_num'],
                connection_fee=request.POST['connection_fee'],
                connection_fee_paid=request.POST['connection_fee_paid'],
                court=court_name,
                network=network,
                email_address=request.POST['email_address'],
                amount_due=int(request.POST['connection_fee'])-int(request.POST['connection_fee_paid'])
                )
        else:
            #cn = 'RB-400'
            court = int(request.POST['court'])

            q = WaterCourt.objects.filter(id=court)

            court_name= q[0].name
            network_id = q[0].network_id
            s = WaterNetwork.objects.filter(id=network_id)
            network = s[0].name





            cn = last_client.client_number
            cn_int = int(cn.split('TNT-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"TNT-{new_cn_int}"
            customer_number = new_cn
            phones = request.POST['msisdn']
            phones2 = request.POST['msisdn2']
            phone_number = f"{0}{phones.replace(' ', '')[-9:]}"
            phone_number2 = f"{0}{phones2.replace(' ', '')[-9:]}"
        WaterClientAll.objects.update_or_create(
            names=request.POST['names'],
            msisdn=phone_number,
            msisdn2=phone_number2,
            client_number=customer_number,
            id_num=request.POST['id_num'],
            connection_fee=request.POST['connection_fee'],
            connection_fee_paid=request.POST['connection_fee_paid'],
            court=court_name,
            network=network,
            email_address=request.POST['email_address'],
            amount_due=int(request.POST['connection_fee'])-int(request.POST['connection_fee_paid'])


        )
        messages.success(request, 'Water Client Added Successfully')
        return redirect('sms:water_clients')
    context = {
        'courts': WaterCourt.objects.filter().order_by('name')
    }
    return render(request, 'sms/add_water_client.html', context)



def create_water_sub_client(request,main_client_id):
    main_client = WaterClientAll.objects.filter(id=main_client_id)
    main_clienter = WaterClientAll.objects.filter(id=main_client_id).first()
    last_client = WaterClientAll.objects.all().order_by('id').last()
    for client in main_client:
        main_client_name  = client.names
        main_client_court=client.court
        main_client_network=client.network
        main_client_ids=client.id
    if request.method == 'POST':
        customer_number = ''
       
        
        #last_client = '1'
        if not last_client:
            customer_number = 'TNT-100'
            cn = 'TNT-100'
            court = int(request.POST['court'])

            q = WaterCourt.objects.filter(id=court)

            court_name= q[0].name
            network_id = q[0].network_id
            s = WaterNetwork.objects.filter(id=network_id)
            network = s[0].name





            #cn = last_client.client_number
            cn_int = int(cn.split('TNT-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"TNT-{new_cn_int}"
            customer_number = new_cn
            phones = request.POST['msisdn']
            provided_account = request.POST['provided_account']
            new_account = main_client_id + '-' + provided_account
            phone_number = f"{0}{phones.replace(' ', '')[-9:]}"
           # phone_number2 = f"{0}{phones2.replace(' ', '')[-9:]}"
            WaterClientAll.objects.update_or_create(
                names=request.POST['names'],
                msisdn=phone_number,
                provided_account=new_account,
                client_number=customer_number,
                court = main_client_network,
                network = main_client_court,
                id_num=request.POST['id_num']
                
                
                )
        else:
            #cn = 'RB-400'
            

            





            cn = last_client.client_number
            cn_int = int(cn.split('TNT-')[-1])
            new_cn_int = cn_int + 1
            new_cn = f"TNT-{new_cn_int}"
            customer_number = new_cn
            phones = request.POST['msisdn']
            provided_account = request.POST['provided_account']
            new_account = str(main_client_id) + '-' + provided_account
            
            phone_number = f"{0}{phones.replace(' ', '')[-9:]}"
            
        WaterClientAll.objects.update_or_create(

            names=request.POST['names'],
            msisdn=phone_number,
            provided_account=new_account.upper(),
            client_number=customer_number,
            court = main_client_network,
            network = main_client_court,
            main_account=main_client_id,
            id_num=request.POST['id_num']
            



        )
        main_clienter.ismain_account=3
        main_clienter.save()
        messages.success(request, 'Water Client Added Successfully')
        return redirect('sms:create_water_sub_client', main_client_id)
    main_client = WaterClientAll.objects.filter(id=main_client_id)
    for client in main_client:
        main_client_name  = client.names
    
    context = {
        'courts': WaterCourt.objects.filter().order_by('name'),
        'main_client_name' : main_client_name,
        'customers': WaterClientAll.objects.filter(main_account=main_client_id),
        'client_id' : main_client_id
    }


    return render(request, 'sms/add_water_sub_client.html', context)


def sample_datatable(request, group_id):
    context = {
        'group': Group.objects.get(id=group_id)
    }
    return render(request, 'contacts/sample.html', context)

def sample_networktable(request, network_id):
    context = {
        'courts': WaterCourt.objects.filter(id=network_id)
    }
    return render(request, 'sms/samplenetwork.html', context)

def sample_water(request, group_id):
    context = {
        'group': Group.objects.get(id=group_id)
    }
    return render(request, 'contacts/sample_water.html', context)


def water_client(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/water_clients.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/water_clients.html', context)

def water_system_config(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        configs = WaterSysConf.objects.filter()
        #WaterSysConf.objects.create(
        #    standing_charge=100,
        #    rate=100
        #)


        context = {
            'configs': configs

        }
        return render(request, 'sms/water_system_config.html', context)


def water_config_history(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        configs = WaterSysConfHist.objects.filter()



        context = {
            'configs': configs

        }
        return render(request, 'sms/water_config_history.html', context)

def add_meter(request):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)

def meter_readingss(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)


def meter_replace(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)


def meter_replacement(request):

    if request.method == 'POST':
        comments = request.POST['comment']
        client_id = request.POST['meter']
        meter_number = request.POST['meter_number']
        customer = WaterClientAll.objects.filter(id=client_id).first()
        last_readings = customer.last_meter_reading
        names = customer.names
        court = customer.court
        amount_due = customer.amount_due
        network=customer.network
        phone_number=customer.msisdn





        WaterMeterReplacement.objects.create(
            client=customer,
            last_units=last_readings,
            comments=comments,
            names=names,
            court=court,
            amount_due=amount_due,
            network=network,
            phone_number=phone_number,
            meter_number=meter_number

        )
        customer.last_meter_reading = 0
        customer.meter_number=meter_number
        customer.save()

        messages.success(request, 'Credit Updated Successfully')
        return redirect('sms:meter_replacement')
    else:
        context = {
            'replacements': WaterMeterReplacement.objects.filter().order_by('-id'),
            'clients': WaterClientAll.objects.filter().order_by('names')
        }
        return render(request, 'sms/meter_replace.html', context)


def main_meter_replacement(request):

    if request.method == 'POST':
        comments = request.POST['comment']
        client_id = request.POST['meter']
        meter_number = request.POST['meter_number']
        customer = WaterNetwork.objects.filter(id=client_id).first()
        #customer = WaterClientAll.objects.filter(id=client_id).first()
        last_readings = customer.reading
        





        MainMeterReplacements.objects.create(
            client=customer,
            last_units=last_readings,
            comments=comments,
            
            meter_number=meter_number

        )
        customer.reading = 0
        customer.meter_num = meter_number
        
        customer.save()

        messages.success(request, 'Meter reset suessfully')
        return redirect('sms:main_meter_replacement')
    else:
        context = {
            'replacements': MainMeterReplacements.objects.filter().order_by('-id'),
            'clients': WaterNetwork.objects.filter().order_by('id')
        }
        return render(request, 'sms/main_meter_replace.html', context)

def water_manual_payments(request):

    if request.method == 'POST':
        comments = request.POST['comment']
        client_id = request.POST['meter']
        amount = request.POST['amount']
        ref_id = request.POST['ref_id']
        customer = WaterClientAll.objects.filter(id=client_id).first()
        names = customer.names

        phone_number=customer.msisdn

        WaterPaymentReceivedManual.objects.create(
            client=customer,
            dest_msisdn=phone_number,
            received_from=names,
            amount=amount,
            confirmation_code=ref_id,
            account_number=client_id,
            account_name=names,
            ref_id=ref_id,
            comments=comments,
            client_id=client_id

        )


        messages.success(request, 'Manual Payment added')
        return redirect('sms:water_manual_payments')
    else:
        context = {
            'payments': WaterPaymentReceivedManual.objects.filter().order_by('-id'),
            'clients': WaterClientAll.objects.filter().order_by('names')
        }
        return render(request, 'sms/water_manual_payment.html', context)

def water_subaccounts_allocations(request,main_id):

    if request.method == 'POST':
        #comments = request.POST['comment']
        client_id = request.POST['client_id']
        new_account_number = request.POST['new_account']

        customer = WaterClientAll.objects.filter(id=client_id).first()
        client = WaterClientAll.objects.filter(id=main_id).first()
        #transaction = MiwamaMpesa.objects.filter(id=trans_id).first()
        customer.main_account=main_id
               
        customer.provided_account = str(main_id) + "-" + new_account_number
        customer.save()
        client.ismain_account=1
        client.save()
        #'clients': WaterClientAll.objects.filter(id__ne=main_id).order_by('names'),
        context = {
            
            'customers': WaterClientAll.objects.filter(main_account=main_id),
            'main_id': main_id
            
        }
        return render(request, 'sms/water_subaccount_allocations.html', context)




        messages.success(request, 'Payments Allocated')
        return redirect('sms:water_payments_allocations')
    else:
        context = {
            
            'clients': WaterClientAll.objects.filter(main_account=0).exclude(id=main_id,ismain_account=1),
            'customers': WaterClientAll.objects.filter(main_account=main_id),
            'main_id': main_id
        }
        return render(request, 'sms/water_subaccount_allocations.html', context)





def water_court_allocations(request):

    if request.method == 'POST':
        comments = request.POST['comment']
        client_id = request.POST['client_id']
        court_id = request.POST['court_id']

        customer = WaterClientAll.objects.filter(id=client_id).first()
        transaction = WaterCourt.objects.filter(id=court_id).first()

       
        prev_court = customer.court
        curr_network = transaction.network.name
        curr_court = transaction.court
        
        new_comment = comments + " Orig court " + prev_court

        customer.court = curr_court
        customer.network = curr_network
        transaction.save()

        WaterCourtReallocate.objects.create(
            client=customer,
            prev_court=prev_court,
            
            curr_network=curr_network,
            comments=new_comment
            

        )


    


        messages.success(request, 'court reallocated')
        return redirect('sms:water_payments_allocations')
    else:
        context = {
            'courts': WaterCourt.objects.filter().order_by('name'),
            'courts_allocated': WaterCourtReallocate.objects.filter().order_by('-id'),
            'clients': WaterClientAll.objects.filter().order_by('names')
        }
        return render(request, 'sms/water_court_allocations.html', context)


def water_payments_allocations(request):

    if request.method == 'POST':
        comments = request.POST['comment']
        client_id = request.POST['client_id']
        trans_id = request.POST['trans_id']

        customer = WaterClientAll.objects.filter(id=client_id).first()
        transaction = MiwamaMpesa.objects.filter(id=trans_id).first()

        names = customer.names
        account_number = customer.id
        phone_number=transaction.sender_phone
        old_account = transaction.account_number
        paid_by=transaction.names
        confirmation_code = transaction.trans_id
        amount=transaction.amount
        new_comment = comments + " Orig account " + old_account

        transaction.processed = 0
        transaction.allocated_to=names
        transaction.is_read=old_account
        transaction.account_number = account_number
        transaction.save()



        WaterPaymentReallocate.objects.create(
            client=customer,
            dest_msisdn=phone_number,
            received_from=paid_by,
            amount=amount,
            confirmation_code=confirmation_code,
            account_number=account_number,
            account_name=names,
            ref_id=trans_id,
            comments=new_comment,
            client_id=client_id

        )


        messages.success(request, 'Payments Allocated')
        return redirect('sms:water_payments_allocations')
    else:
        context = {
            'payments': MiwamaMpesa.objects.filter(processed=3).order_by('-id'),
            'payments_allocated': WaterPaymentReallocate.objects.filter().order_by('-id'),
            'clients': WaterClientAll.objects.filter().order_by('names')
        }
        return render(request, 'sms/water_payment_allocations.html', context)

def meter_replacemet2(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        #cleints = WaterClientsAll.objects.all(customer=customer.id)

        context = {
            'clients': WaterClientAll.objects.filter().order_by('names')
        }
        return render(request, 'sms/meter_replacement.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'clients': WaterClientAll.objects.filter().order_by('-id')
        }
        return render(request, 'sms/contacts.html', context)
def roberms_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            if Customer.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('sms:water_apps')
            elif SalesPerson.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('salesperson:dashboard')
            elif Manager.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('roberms_admin:dashboard')
            elif CustomerSubAccounts.objects.filter(user_ptr_id=user.id).count() > 0:
                print(user.id)
                return redirect('sms:water_apps')
        else:
            messages.error(request, 'Invalid Email Or Password')
            return redirect('sms:login')
    return render(request, 'registration/login.html')


def create_water_network(request):
    customer = Customer.objects.filter(id=request.user.id).first()
    if customer is not None:
        if request.method == "POST":
            WaterNetwork.objects.create(
                customer=customer,
                name=request.POST['name'],
                rate=request.POST['rate'],
                standing_charge=request.POST['standing_charge']
            )
            return redirect('sms:water_courts')
        return render(request, 'sms/create_network.html')
    else:
        customer = CustomerSubAccounts.objects.filter(id=request.user.id).first().owner
        if request.method == "POST":
            Group.objects.create(
                customer_id=customer,
                name=request.POST['name'],
                standing_charge=request.POST['standing_charge'],
                rate=request.POST['rate']
            )
            return redirect('sms:water_courts')
        return render(request, 'group/create_network.html')

@login_required()
@is_user_customer
def create_group(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:customer_contacts')
        return render(request, 'group/create_group.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:customer_contacts')
        return render(request, 'group/create_group.html')

def create_network(request):
    customer = Customer.objects.filter(id=request.user.id).first()
    if customer is not None:
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:water_courts')
        return render(request, 'sms/create_network.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:customer_contacts')
        return render(request, 'group/create_group.html')
@login_required()
@is_user_customer
def group_contacts(request, group_id):
    contact_list = Contact.objects.filter(group_id=group_id)
    # paginator = Paginator(contact_list, 500)
    # page = request.GET.get('page', 1)

    # try:
    #     contacts = paginator.page(page)
    # except PageNotAnInteger:
    #     contacts = paginator.page(1)
    # except EmptyPage:
    #     contacts = paginator.page(paginator.num_pages)

    context = {
        'contacts': contact_list,
        'group': Group.objects.filter(id=group_id).first(),
        # 'record_count': paginator.count
    }
    return render(request, 'contacts/group_contacts.html', context)


def water_courts_network(request, network_id):
    court_list = WaterCourt.objects.filter(network_id=network_id)
    # paginator = Paginator(contact_list, 500)
    # page = request.GET.get('page', 1)

    # try:
    #     contacts = paginator.page(page)
    # except PageNotAnInteger:
    #     contacts = paginator.page(1)
    # except EmptyPage:
    #     contacts = paginator.page(paginator.num_pages)

    context = {

        'coourts': court_list,
        'network': WaterNetwork.objects.filter(id=network_id).first(),
    }
    return render(request, 'sms/samplenetwork.html', context)

@login_required()
@is_user_customer
def activate_deactivate_contact(request, contact_id):
    contact = Contact.objects.get(id=contact_id)
    if contact.is_active:
        contact.is_active = False
        contact.save()
        messages.success(request, 'Contact deactivated successfully')
        return redirect('sms:update_contact', contact.id)
    elif not contact.is_active:
        contact.is_active = True
        contact.save()
        messages.success(request, 'Contact activated successfully')
        return redirect('sms:update_contact', contact.id)


@login_required()
@is_user_customer
def create_contact(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    if request.method == 'POST':
        phone_number = f"{254}{request.POST['phone_number'].replace(' ', '')[-9:]}"
        Contact.objects.update_or_create(
            group_id=group_id,
            name=request.POST['name'],
            email=request.POST['email'],
            phone_number=phone_number
        )
        return redirect('sms:sample_datatable', group_id)
    context = {
        'group': group
    }
    return render(request, 'contacts/create_contact.html', context)
def create_court(request, network_id):
    network = WaterNetwork.objects.filter(id=network_id).first()
    if request.method == 'POST':
        #phone_number = f"{254}{request.POST['phone_number'].replace(' ', '')[-9:]}"
        WaterCourt.objects.create(
            network_id=network_id,
            name=request.POST['name'],
            location=request.POST['location']

        )
        return redirect('sms:sample_datatable_network', network_id)
    context = {
        'network': network
    }
    return render(request, 'sms/create_network_court.html', context)
@timed
@login_required()
@is_user_customer
def import_contacts(request, group_id):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)

def import_water_clients(request):
    if request.method == 'POST':
        last_client = WaterClientAll.objects.all().order_by('id').last()
        #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        customer = Customer.objects.filter(id=1).first()


        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            #f_path = "templates"
            extension = file.name.rsplit('.', 1)[1]
            workbook = load_workbook(filename=f_path, read_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            for i in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=i, column=1).value != '':

                    names = worksheet.cell(row=i, column=1).value
                    phone_numbers = str(worksheet.cell(row=i, column=2).value)
                    phone_numbers2 = str(worksheet.cell(row=i, column=3).value)
                    id_number = (worksheet.cell(row=i, column=4).value)
                    meter_number = (worksheet.cell(row=i, column=5).value)
                    court = (worksheet.cell(row=i, column=6).value)
                    meter_readings = (worksheet.cell(row=i, column=7).value)
                    amount_due = float(worksheet.cell(row=i, column=8).value)
                    connection_fee = worksheet.cell(row=i, column=9).value
                    connection_fee_paid = worksheet.cell(row=i, column=10).value
                    email = worksheet.cell(row=i, column=11).value
                    phone_number = f"{0}{phone_numbers.replace(' ', '')[-9:]}"
                    phone_number2 = f"{0}{phone_numbers2.replace(' ', '')[-9:]}"



                    q = WaterCourt.objects.filter(name=court)

                    court_name = q[0].name
                    network_id = q[0].network_id
                    s = WaterNetwork.objects.filter(id=network_id)
                    network = s[0].name

                    cn = last_client.client_number
                    cn_int = int(cn.split('TNT-')[-1])
                    new_cn_int = cn_int + 1
                    new_cn = f"TNT-{new_cn_int}"
                    customer_number = new_cn
                if not WaterClientAll.objects.filter(names=names).exists():
                    WaterClientAll.objects.update_or_create(
                        names=names,
                        msisdn=phone_number,
                        msisdn2=phone_number2,
                        client_number=customer_number,
                        id_num=id_number,
                        meter_num=meter_number,
                        connection_fee=connection_fee,
                        connection_fee_paid=connection_fee_paid,
                        last_meter_reading=meter_readings,

                        court=court,
                        network=network,

                        email_address=email,
                        amount_due=amount_due

                    )




            messages.success(request, filename)

            context = {
                'clients': WaterClientAll.objects.filter().order_by('-id')
            }
            return render(request, 'sms/water_clients.html', context)

        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_meter_readings_task.delay(extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)

def import_water_sub_clients(request, client_id):
    if request.method == 'POST':
        main_client = WaterClientAll.objects.filter(id=client_id)
        main_clienter = WaterClientAll.objects.filter(id=client_id).first()
        last_client = WaterClientAll.objects.all().order_by('id').last()

        for client in main_client:
            
            main_client_name  = client.names
            main_client_court=client.court
            main_client_network=client.network
        #customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        customer = Customer.objects.filter(id=1).first()


        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            #f_path = "templates"
            extension = file.name.rsplit('.', 1)[1]
            workbook = load_workbook(filename=f_path, read_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            for i in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=i, column=1).value != '':

                    names = worksheet.cell(row=i, column=1).value
                    phone_numbers = str(worksheet.cell(row=i, column=2).value)
                   
                    id_number = (worksheet.cell(row=i, column=3).value)
                    new_provided_account = (worksheet.cell(row=i, column=4).value)
                    court = (worksheet.cell(row=i, column=6).value)
                    
                    
                    phone_number = f"{0}{phone_numbers.replace(' ', '')[-9:]}"
                    



                   

                    cn = last_client.client_number
                    cn_int = int(cn.split('TNT-')[-1])
                    new_cn_int = cn_int + 1
                    new_cn = f"TNT-{new_cn_int}"
                    customer_number = new_cn
                if not WaterClientAll.objects.filter(names=names).exists():
                    WaterClientAll.objects.update_or_create(
                        names=names,
                        msisdn=phone_number,
                        
                        client_number=customer_number,
                        id_num=id_number,
                        provided_account = (str(client_id) + "-" + new_provided_account).upper(),
                        
                        main_account=client_id,
                        

                        court=main_client_court,
                        network=main_client_network

                        
                        

                    )
                main_clienter.ismain_account=1
                main_clienter.save()
            




            messages.success(request, filename)

            context = {
                'clients': WaterClientAll.objects.filter().order_by('-id')
            }
            return redirect('sms:create_water_sub_client', client_id)

        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_meter_readings_task.delay(extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
def import_water_clientss(request):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)

@task()
def store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    Contact.objects.update_or_create(
                        name=row[0],
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
            CustomerTask.objects.filter(task_id=store_contact_task.id).update(
                status_complete=True
            )
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        for i in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=i, column=2).value != '':
                group_id = group_id
                name = worksheet.cell(row=i, column=1).value
                phone_number = str(worksheet.cell(row=i, column=2).value)
                email = worksheet.cell(row=i, column=3).value
                p = f"{254}{phone_number.replace(' ', '')[-9:]}"
                try:
                    Contact.objects.update_or_create(
                        name=name,
                        group_id=group_id,
                        phone_number=int(p),
                        email=email
                    )
                except TypeError:
                    continue
                except ValueError:
                    continue
                print('saved')
        # CustomerTask.objects.filter(task_id=store_contact_task.id).update(
        #     status_complete=True
        # )
        return 'completed'


@timed
@login_required()
@is_user_customer
def express_import_contacts(request, group_id):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = new_store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = new_store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)


@task()
def new_store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            contacts = []
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    contact = Contact(
                        name=str(row[0]),
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
                    contacts.append(contact)
                    if len(contacts) >= 20000:
                        Contact.objects.bulk_create(contacts)
                        contacts.clear()
            Contact.objects.bulk_create(contacts)
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=False)
        worksheet = workbook[workbook.sheetnames[0]]
        contacts = []
        for i in range(2, worksheet.max_row):
            # if worksheet.cell(row=i, column=2).value != '':
            group_id = group_id
            name = worksheet.cell(row=i, column=1).value
            phone_number = str(worksheet.cell(row=i, column=2).value)
            email = worksheet.cell(row=i, column=3).value
            p = f"{254}{phone_number.replace(' ', '')[-9:]}"

            # Contact.objects.update_or_create(
            #     name=name,
            #     group_id=group_id,
            #     phone_number=int(p),
            #     email=email
            # )
            contact = Contact(
                name=name,
                group_id=group_id,
                phone_number=int(p),
                email=email
            )
            contacts.append(contact)
            if len(contacts) >= 20000:
                Contact.objects.bulk_create(contacts)
                contacts.clear()
            else:
                continue
        Contact.objects.bulk_create(contacts)
        return 'completed'


@login_required()
@is_user_customer
def contacts_upload_status(request, task_id):
    context = {
        'task_id': task_id
    }
    return render(request, 'contacts/upload_status.html', context)


def poll_contact_upload_state(request, task_id):
    """ A view to report the progress to the user """

    job = AsyncResult(task_id)
    data = job.result or job.state
    response_data = {
        'state': job.state,
        'details': job.result,
    }
    return HttpResponse(json.dumps(response_data), content_type='application/json')


@login_required()
@is_user_customer
def comprehensive_reports(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        # messages_sent = OutgoingDone.objects.filter(customer_id=customer.id).values('track_code', 'sent_time').distinct()
        messages_sent = OutgoingDone.objects.raw(
            f'SELECT * FROM sms_outgoingdone WHERE customer_id={customer.id} GROUP BY track_code'
        )
        track_codes = []
        # for message in messages_sent.iterator():
        #     track_codes.append(message.track_code)

        print(messages_sent)
        context = {
            'track_codes': messages_sent
        }
        return render(request, 'sms/comprehensive_reports.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        # messages_sent = OutgoingDone.objects.filter(customer_id=customer.id).values('track_code', 'sent_time').distinct()
        messages_sent = OutgoingDone.objects.raw(
            f'SELECT * FROM sms_outgoingdone WHERE customer_id={customer.id} GROUP BY track_code'
        )
        # track_codes = []
        # for message in messages_sent.iterator():
        #     track_codes.append(message.track_code)
        context = {
            'track_codes': messages_sent
        }
        return render(request, 'sms/comprehensive_reports.html', context)




@login_required()
@is_user_customer
def report_details(request, track_code):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        o_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id)
        sent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='DeliveredTo')
        unsent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id).exclude(
            delivery_status__contains='DeliveredTo')
        pending_delivery_status = OutgoingNew.objects.filter(track_code=track_code, customer_id=customer.id,
                                                             delivery_status__isnull=True)
        context = {
            # 'o_messages': o_messages,
            'sent_messages': sent_messages.count(),
            'unsent_messages': unsent_messages.count(),
            'pending_delivery_status': pending_delivery_status.count(),
            'track_code': track_code,
            'airtel_messages': OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='success').count()
        }
        return render(request, 'sms/report_details.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        # o_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id)
        sent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='DeliveredTo')
        unsent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id).exclude(
            delivery_status__contains='DeliveredTo')
        pending_delivery_status = OutgoingNew.objects.filter(track_code=track_code, customer_id=customer.id,
                                                             delivery_status__isnull=True)

        context = {
            # 'o_messages': o_messages,
            'sent_messages': sent_messages.count(),
            'unsent_messages': unsent_messages.count(),
            'pending_delivery_status': pending_delivery_status.count(),
            'track_code': track_code,
            'airtel_messages': OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                           delivery_status__contains='success').count()
        }
        return render(request, 'sms/report_details.html', context)


# @login_required()
# @is_user_customer
# def generate_sms_report(request, track_code):
#     messages = OutgoingDone.objects.filter(track_code=track_code)
#     customer = Customer.objects.filter(user_ptr_id=messages.first().customer_id).first()
#     time = datetime.datetime.now()
#     file_path = 'media/reports/%s' % (customer.sender_name.replace(" ", "_"))
#     filename = "%s_%d_%d.pdf" % (messages.first().track_code, time.year, time.month)
#     full_path = f"{file_path}/{filename}"
#     data = {
#         'today': datetime.datetime.today(),
#         'messages': messages
#     }
#
#     html_string = render_to_string('sms/sms_report.html', data)
#
#     html = HTML(string=html_string)
#     html.write_pdf(target=full_path)
#
#     fs = FileSystemStorage(file_path)
#     with fs.open(filename) as pdf:
#         print(pdf)
#         response = HttpResponse(pdf, content_type='application/pdf')
#         # response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         return response


@login_required()
@is_user_customer
def personalized_sms_menu(request):
    return render(request, 'sms/personalized_sms_menu.html')


@login_required()
@is_user_customer
def personalized_from_contact_list(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            group_ids = [g.id for g in groups]
            if 'all_groups' in request.POST:
                context = {
                    'message': message,
                    'groups': group_ids
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
            else:
                context = {
                    'message': message,
                    'group': request.POST['group']
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
        context = {
            'groups': groups
        }
        return render(request, 'sms/personalized_from_contact_list.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            group_ids = [g.id for g in groups]
            if 'all_groups' in request.POST:
                context = {
                    'message': message,
                    'groups': group_ids
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
            else:
                context = {
                    'message': message,
                    'group': request.POST['group']
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
        context = {
            'groups': groups
        }
        return render(request, 'sms/personalized_from_contact_list.html', context)


@login_required()
@is_user_customer
def send_to_plain(request):
    import time
    t = time.localtime()
    current_time = time.strftime("%Y%m%d", t)
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()

    if customer is not None:
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()

            for phone_number in phone_numbers:
                if phone_number != '':

                    #data = {}
                    # print(phone_numbers)
                    new_phone_numbers = []
                    for phone_number in phone_numbers:
                        if len(phone_number) > 8:
                            new_phone_numbers.append(phone_number)

                    for p in new_phone_numbers:
                        phone_number = f"{0}{p[-9:]}"
                        WaterOutbox.objects.create(
                            user_id=customer.id,
                            client=customer.id,
                            text_message=message,
                            dest_msisdn=phone_number,
                            track_code=current_time,
                            delivery_status='DeliveredToNet',
                            timestamp=timezone.now()
                        )


                return redirect("sms:water_sms_reports", current_time)

        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_sms.html', context)


@login_required()
@is_user_customer
def send_to_all(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            msg_len = len(message)
            WaterSmsOut.objects.create(
                user_id=customer.id,
                client=customer.id,
                text_message=message,
                message_length=msg_len,

                timestamp=timezone.now()
            )
            last_id = WaterSmsOut.objects.filter().last().id




            return redirect('sms:customer_reports', last_id)

        return render(request, 'sms/send_to_all.html')


@login_required()
@is_user_customer
def send_to_court(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        #groups = WaterClientAll.objects.all().order_by('court').distinct()
        #groups = WaterCourt.objects.all().order_by('name').distinct()
        #groups = WaterCourt.objects.all().values('name').distinct()
        groups = WaterClientAll.objects.all().values('court').distinct()
        #groups = WaterClientAll.objects.order_by().values_list('court',Flat=True).distinct()
        if request.method == 'POST':
            message = request.POST.get('Message')
            court = request.POST.get('group')
            msg_len = len(message)
            WaterSmsOut.objects.create(
                user_id=customer.id,
                client=customer.id,
                text_message=message,
                message_length=msg_len,
                court=court,

                timestamp=timezone.now()
            )
            last_id = WaterSmsOut.objects.filter().last().id






            return redirect('sms:customer_reports', last_id)
        context = {
            'groups': groups
        }

        return render(request, 'sms/send_to_court.html',context)

@login_required()
@is_user_customer
def send_to_network(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = WaterClientAll.objects.all().values('network').distinct()
        #groups = WaterNetwork.objects.all().values('name').distinct()

        #groups = WaterClientAll.objects.order_by().values_list('court',Flat=True).distinct()
        if request.method == 'POST':
            message = request.POST.get('Message')
            network = request.POST.get('group')
            msg_len = len(message)
            WaterSmsOut.objects.create(
                user_id=customer.id,
                client=customer.id,
                text_message=message,
                message_length=msg_len,
                network=network,

                timestamp=timezone.now()
            )
            last_id = WaterSmsOut.objects.filter().last().id






            return redirect('sms:customer_reports', last_id)
        context = {
            'groups': groups
        }

        return render(request, 'sms/send_to_network.html',context)
@login_required()
@is_user_customer
def send_reminder_network(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = WaterClientAll.objects.all().values('network').distinct()
        #groups = WaterNetwork.objects.all().values('name').distinct()

        #groups = WaterClientAll.objects.order_by().values_list('court',Flat=True).distinct()
        if request.method == 'POST':
            message = "Regards"
            network = request.POST.get('group')
            msg_len = len(message)
            WaterSmsOutArrearsReminder.objects.create(
                user_id=customer.id,
                client=customer.id,
                text_message=message,
                message_length=msg_len,
                network=network,

                timestamp=timezone.now()
            )
            last_id = WaterSmsOutArrearsReminder.objects.filter().last().id






            return redirect('sms:customer_reports', last_id)
        context = {
            'groups': groups
        }

        return render(request, 'sms/send_reminder_network.html',context)


def c_sample_merged(request):
    data = request.session.get('c_data')
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        parameters = get_message_parameters(message=data['message'])
        if 'group' in data:
            c_group = Group.objects.filter(id=data['group']).first()
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for contact in Contact.objects.filter(group_id=request.POST['group'], is_active=True):
                    complete_message = message
                    if parameters:
                        for parameter in parameters:
                            if parameter == 'Name':
                                complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                    else:
                        to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'group':
                        contacts = Contact.objects.filter(group_id=b)
                        for contact in contacts:
                            complete_message = data['message']
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': c_group.id
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
        elif 'groups' in data:
            parameters = get_message_parameters(message=data['message'])
            groups = Group.objects.filter(id__in=data['groups'])

            if request.method == 'POST':
                print('got here')
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']
                print(groups)
                for group in groups:
                    for contact in Contact.objects.filter(group_id=group.id, is_active=True):
                        complete_message = message
                        if parameters:
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                        else:
                            to_be_sent[contact.phone_number] = complete_message
                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'groups':
                        for g in groups:
                            contacts = Contact.objects.filter(group_id=g.id)
                            for contact in contacts:
                                complete_message = data['message']
                                for parameter in parameters:
                                    if parameter == 'Name':
                                        complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': groups[0]
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        parameters = get_message_parameters(message=data['message'])
        if 'group' in data:
            c_group = Group.objects.filter(id=data['group']).first()
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for contact in Contact.objects.filter(group_id=request.POST['group'], is_active=True):
                    complete_message = message
                    if parameters:
                        for parameter in parameters:
                            if parameter == 'Name':
                                complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                    else:
                        to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'group':
                        contacts = Contact.objects.filter(group_id=b)
                        for contact in contacts:
                            complete_message = data['message']
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': c_group.id
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
        elif 'groups' in data:
            groups = Group.objects.filter(id__in=data['groups'])
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for group in groups:
                    for contact in Contact.objects.filter(group_id=group.id, is_active=True):
                        complete_message = message
                        if parameters:
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                        else:
                            to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'groups':
                        for g in groups:
                            contacts = Contact.objects.filter(group_id=g.id)
                            for contact in contacts:
                                complete_message = data['message']
                                for parameter in parameters:
                                    if parameter == 'Name':
                                        complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': groups[0]
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)


@task()
def from_group_send(customer_id, total_message_cost, to_be_sent, trackingcode):
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    # messages = []
    for a, b in to_be_sent.items():
        outgoing_new, created = OutgoingNew.objects.update_or_create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=a,
            text_message=b,
            track_code=trackingcode,
            sent_time=timezone.now()
        )

    #     messages.append(
    #         OutgoingNew(
    #             customer=customer,
    #             service_id=customer.service_id,
    #             access_code=customer.access_code,
    #             phone_number=a,
    #             text_message=b,
    #             track_code=trackingcode
    #         )
    #     )
    #     if len(messages) > 1000:
    #         OutgoingNew.objects.bulk_create(messages)
    # if len(messages) > 0:
    #     OutgoingNew.objects.bulk_create(messages)
    send_usage(total_message_cost)
    return 'completed insertion'


@login_required
def delete_group(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    group.delete()
    messages.success(request, 'Group Successfully Deleted')
    return redirect("sms:customer_contacts")


@login_required
def delete_contact(request, contact_id):
    contact = Contact.objects.filter(id=contact_id).first()
    group = contact.group
    contact.delete()
    messages.success(request, 'Contact Successfully Deleted')
    return redirect("sms:sample_datatable", group.id)


def customer_top_up(request):
    tracking_code = random.randint(1, 1000000)
    sender_phone = request.POST['sender_phone']
    transaction_reference = request.POST['trans_id']
    amount = request.POST['amount']
    till_number = request.POST['till_number']
    timestamp = request.POST['transaction_timestamp']
    first_name = request.POST['first_name']
    last_name = request.POST['last_name']
    # transaction_type = request.POST['transaction_type']

    UserTopUp.objects.create(
        phone_number=sender_phone,
        transaction_ref=transaction_reference,
        amount=amount,
        till_number=till_number,
        f_name=first_name,
        l_name=last_name,
        verify_code=tracking_code,
        timestamp=timestamp
    )

    message = f"Thank you {first_name}" \
        "for paying for Roberms sms service. To automatically load the credit into your account, " \
        f"use this code to verify your payment. {tracking_code}"
    customer = Customer.objects.filter().first()
    if customer is not None:
        # if str.isdigit(customer.access_code):
        #     outgoing = Outgoing.objects.create(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=sender_phone,
        #         text_message=message,
        #         track_code=tracking_code
        #     )
        # else:
        outgoing_new = OutgoingNew.objects.create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=sender_phone,
            text_message=message,
            track_code=tracking_code,
            sent_time=timezone.now()
        )
    else:
        customer = CustomerSubAccounts.objects.filter().first().owner
        # if str.isdigit(customer.access_code):
        #     outgoing = Outgoing.objects.create(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=sender_phone,
        #         text_message=message,
        #         track_code=tracking_code
        #     )
        # else:
        outgoing_new = OutgoingNew.objects.create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=sender_phone,
            text_message=message,
            track_code=tracking_code,
            sent_time=timezone.now()
        )
        # sdp = SDP()
        # response = sdp.send_sms_customized(service_id=outgoing.service_id, recipients=[outgoing.phone_number],
        #                                    message=outgoing.text_message, sender_code='')


@login_required()
@is_user_customer
def verify_payment(request):
    if request.method == 'POST':
        user_top_up = Sms_TopUp.objects.filter(verifycode=request.POST['verification_code'], verified=0).first()
        if user_top_up:
            credit = 0
            if float(user_top_up.amount) < 500:
                credit = float(user_top_up.amount) * float(1)
            elif 500 <= float(user_top_up.amount) <= float(9999):
                credit = float(user_top_up.amount) * float(1.67)
            elif 10000 <= float(user_top_up.amount) <= float(49999):
                credit = float(user_top_up.amount) * float(2)
            elif float(user_top_up.amount) >= float(50000):
                credit = float(user_top_up.amount) * float(2.5)

            customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
            if customer is not None:
                existing_credit = customer.credit
                customer.credit = existing_credit + credit
                user_top_up.verified = 1
                user_top_up.user_id = customer.id
                user_top_up.created_at = datetime.datetime.now()
                user_top_up.save()
                customer.save()
                ManagerTopUp.objects.create(
                    sms_count=credit,
                    amount=user_top_up.amount,
                    user_id=customer.id,
                    timestamp=datetime.datetime.now()
                )
                messages.success(request, 'Top Up Successful')
                return redirect('sms:apps')
            else:
                customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
                existing_credit = customer.credit
                customer.credit = existing_credit + credit
                user_top_up.verified = 1
                user_top_up.user_id = customer.id
                user_top_up.created_at = datetime.datetime.now()
                user_top_up.save()
                customer.save()
                ManagerTopUp.objects.create(
                    sms_count=credit,
                    amount=user_top_up.amount,
                    user_id=customer.id,
                    timestamp=datetime.datetime.now()
                )
                messages.success(request, 'Top Up Successful')
                return redirect('sms:apps')
        else:
            messages.error(request, 'The Verification Code You Entered is Invalid')
            return redirect('sms:verify_payment')
    return render(request, 'top_up/payment_verification_code.html')


@login_required
def update_group(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    if request.method == 'POST':
        Group.objects.filter(id=group_id).update(
            name=request.POST['name']
        )
        messages.success(request, 'Group Successfully Updated')
        return redirect('sms:customer_contacts')
    context = {
        'group':  group
    }
    return render(request, 'sms/update_group.html', context)

def update_network(request, group_id):
    network = WaterNetwork.objects.filter(id=network_id).first()
    if request.method == 'POST':
        WaterNetwork.objects.filter(id=network_id).update(
            name=request.POST['name']
        )
        messages.success(request, 'Network Successfully Updated')
        return redirect('sms:water_courts')
    context = {
        'networks':  network
    }
    return render(request, 'sms/update_network.html', context)

@login_required
def update_contact(request, contact_id):
    contact = Contact.objects.filter(id=contact_id).first()
    if request.method == 'POST':
        Contact.objects.filter(id=contact_id).update(
            name=request.POST['name'],
            email=request.POST['email'],
            phone_number=request.POST['phone_number']
        )
        messages.success(request, 'Contact Successfully Updated')
        return redirect('sms:group_contacts', contact.group.id)
    context = {
        'group': contact.group,
        'contact': contact
    }
    return render(request, 'sms/update_contact.html', context)


@login_required
def sms_reports(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')

def simple_sms_resend(request, message_id):
    client = WaterOutbox.objects.get(id=message_id)
    if request.method == 'POST':
       
        WaterOutbox.objects.create(
            dest_msisdn=request.POST['dest_msisdn'],
            text_message=request.POST['text_message'],
            user_id=100,
            client=message_id
            


        )

        messages.success(request, "Message resent sucessfully")
        return redirect('sms:water_sent_sms')
    context = {
        'client': client
    }
    return render(request, 'sms/resend_message.html', context)

@login_required
def water_sent_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        #m = WaterOutbox.objects.filter().order_by('-id').values()
        m = WaterOutbox.objects.filter().order_by('-out_date')[:600]
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/water_sent_sms.html', context)
        else:
            return render(request, 'sms/water_sent_sms.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')

@login_required
def water_sent_sms_client(request,client_phone):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        m = WaterOutbox.objects.filter(dest_msisdn=client_phone)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/water_sent_sms.html', context)
        else:
            return render(request, 'sms/water_sent_sms.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')

@login_required()
def customer_till_numbers(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        till_numbers = Till_Numbers.objects.filter(customer=customer)
        context = {
            'till_numbers': till_numbers
        }
        return render(request, 'sms/customer_till_numbers.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        till_numbers = Till_Numbers.objects.filter(customer=customer)
        context = {
            'till_numbers': till_numbers
        }
        return render(request, 'sms/customer_till_numbers.html', context)


@login_required()
def delete_till_number(request, till_number_id):
    till_number = Till_Numbers.objects.filter(id=till_number_id).first()
    if till_number is not None:
        till_number.delete()
        messages.success(request, 'Till Number Deleted Successfully')
        return redirect('sms:customer_till_numbers')
    else:
        messages.success(request, 'Till Number Does Not Exist')
        return redirect('sms:customer_till_numbers')


@login_required()
def edit_till_number(request, till_number_id):
    till_number = Till_Numbers.objects.get(id=till_number_id)
    if request.method == 'POST':
        till_number = request.POST['till_number']
        message = request.POST['message']

        Till_Numbers.objects.filter(id=till_number_id).update(
            till=till_number,
            message=message
        )

        return redirect('sms:customer_till_numbers')
    else:
        context = {
            'till_number': till_number
        }
        return render(request, 'till_numbers/edit_till_number.html', context)

#
@login_required()
def add_till_number(request):
    if request.method == 'POST':
        till_number = request.POST['till_number']
        message = request.POST['message']
        customer = customer=Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            Till_Numbers.objects.create(
                customer=customer,
                till=till_number,
                message=message
            )
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            if customer is not None:
                Till_Numbers.objects.create(
                    customer=customer,
                    till=till_number,
                    message=message
                )
        return redirect('sms:customer_till_numbers')
    return render(request, 'till_numbers/add_till_number.html')


@login_required()
def delete_till_number(request, till_number_id):
    till = Till_Numbers.objects.filter(id=till_number_id).first()
    if till is not None:
        till.delete()
        messages.success(request, 'Till Number Deleted')
        return redirect('sms:customer_till_numbers')
    messages.error('Unable to delete till number')
    return redirect('sms:customer_till_numbers')


@login_required()
def credit_used(request):
    months = get_last_n_months(6)
    monthly_credit = []
    for month in list(set(months)):
        messages = OutgoingDone.objects.filter(sent_time__month=month)
        credit = 0
        for message in messages:
            credit += calculate_message_cost(message.text_message)
        monthly_credit.append(credit)

    context = {
        'months': list(set(months)),
        'monthly_credit': monthly_credit,
    }
    return render(request, 'sms/credit_usage.html', context)


@login_required()
def applications(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        applications = SenderNameApplication.objects.filter(customer_id=customer.id)
        context = {
            'applications': applications
        }
        return render(request, 'sender_name/sender_name_list.html', context)
    else:
        sub_account = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first()
        customer = sub_account.owner
        applications = SenderNameApplication.objects.filter(customer_id=customer.id)
        context = {
            'applications': applications
        }
        return render(request, 'sender_name/sender_name_list.html', context)


@login_required()
def new_application(request):
    if request.method == 'POST':
        date = datetime.datetime.today().date()
        sender_name = request.POST['sender_name']
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            n_application = SenderNameApplication.objects.create(
                customer=customer,
                sender_name=sender_name,
                application_date=date,
            )
            return redirect('sms:application_contacts', n_application.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            n_application = SenderNameApplication.objects.create(
                customer=customer,
                sender_name=sender_name,
                application_date=date,
            )
            return redirect('sms:application_contacts', n_application.id)
    else:
        return render(request, 'sender_name/application_form.html')


@login_required()
def application_contacts(request, application_id):
    application_contacts = ApplicationContact.objects.filter(application_id=application_id)
    context = {
        'contacts': application_contacts,
        'application_id': application_id
    }
    return render(request, 'sender_name/contacts.html', context)


@login_required()
def add_application_contacts(request, application_id):
    application = SenderNameApplication.objects.filter(id=application_id).first()
    if request.method == 'POST':
        name = request.POST['name']
        phone_number = request.POST['phone_number']
        application_id = request.POST['application_id']
        ApplicationContact.objects.create(
            name=name,
            phone_number=phone_number,
            application_id=application_id
        )
        return redirect('sms:application_contacts', application_id)
    else:
        return render(request, 'sender_name/add_contacts.html', {'application':application})


@login_required()
def show_pdf(request, application_id):
    application = SenderNameApplication.objects.filter(id=application_id).first()
    application_contacts = ApplicationContact.objects.filter(application=application)
    data = {
        'application': application,
        'contacts': application_contacts
    }
    pdf = render_to_pdf('sms/sender_name_application.html', data)
    return HttpResponse(pdf, content_type='application/pdf')


# @login_required()
# def inbox(request):
#     customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
#     if customer is not None:
#         inbox_messages = Inbox.objects.filter(customer_id=customer.id)
#         context = {
#             'messages': inbox_messages
#         }
#         return render(request, 'inbox/general_inbox.html', context)
#     else:
#         customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
#         inbox_messages = Inbox.objects.filter(customer_id=customer.id)
#         context = {
#             'messages': inbox_messages
#         }
#         return render(request, 'inbox/general_inbox.html', context)


def airtel_callback(request):
    print(request)
    return


def trial(request):
    print(get_access_token())
    return HttpResponse(get_access_token())


def documentation(request):
    return render(request, 'api/documentation.html')


def offers(request):

    return render(request, 'sms/offers.html')


@login_required()
def new_tag(request):
    if request.method == "POST":
        tag = request.POST['tag']
        response = request.POST['default_response']
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            tag = f'#{tag}'
            tag = Tag.objects.update_or_create(
                hashtag=tag, customer=customer, defaults={'response': response}
            )
            messages.success(request, 'Tag created successfully')
            return redirect('sms:get_tags')
        else:
            tag = f'#{tag}'
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            tag, created = Tag.objects.update_or_create(
                hashtag=tag, customer=customer, defaults={'response': response}
            )
            messages.success(request, 'Tag created successfully')
            return redirect('sms:get_tags')
    else:
        return render(request, 'inbox/new_tag.html')


def get_tags(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        context = {
            'tags': Tag.objects.filter(customer_id=customer.id)
        }
        return render(request, 'inbox/tags.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        context = {
            'tags': Tag.objects.filter(customer_id=customer.id)
        }
        return render(request, 'inbox/tags.html', context)


def tag_messages(request, tag_id):
    context = {
        'tag_messages': Inbox.objects.filter(tag_id=tag_id)
    }
    return render(request, 'inbox/general_inbox.html', context)


def temp_clean_sys(request):
    outgoings = OutgoingDone.objects.filter(
        Q(delivery_status="SenderName Blacklisted")|
        Q(delivery_status="InvalidMsisdn")|
        Q(delivery_status="AbsentSubscriber")
        ,
        Q(access_code="TREASURE")|
        Q(access_code="Nyumbani"),
    )
    print(outgoings.query)

    groups = Group.objects.filter(
        Q(customer_id=148) |
        Q(customer_id=52)
    )

    for outgoing in outgoings:
        for group in groups:
            contact = Contact.objects.filter(phone_number=outgoing.phone_number, group_id=group.id).first()
            if contact is not None:
                contact.is_active = False
                contact.save()
    return redirect("sms:apps")


from django_datatables_view.base_datatable_view import BaseDatatableView
from django.utils.html import escape


class OrderListJson(BaseDatatableView):
    model = Contact
    columns = ['id', 'name', 'phone_number', 'email', 'creation_date', 'is_active', 'actions']
    order_columns = ['id', 'name', 'phone_number', 'email', 'creation_date', '']
    max_display_length = 200

    def get_initial_queryset(self):
        print(self.kwargs['group_id'])
        return self.model.objects.filter(group_id=self.kwargs['group_id']).\
            extra(select={'creation_date'})

    def render_column(self, row, column):
        if column == 'user':
            return escape('{0} {1}'.format(row.customer_firstname, row.customer_lastname))
        else:
            return super(OrderListJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]', None)
        print(self.kwargs['group_id'])
        if search:
            qs = qs.filter(Q(name__istartswith=search)|Q(phone_number__istartswith=search), group_id=self.kwargs['group_id'])

        filter_customer = self.request.GET.get('customer', None)
        # print(filter_customer)
        if filter_customer:
            customer_parts = filter_customer.split(' ')
            qs_params = None
            for part in customer_parts:
                q = Q(name__istartswith=part)|Q(phone_number__istartswith=part)
                qs_params = qs_params | q if qs_params else q
            qs = qs.filter(qs_params)
        return qs







class OrderListJsonNet(BaseDatatableView):
    model = WaterCourt
    columns = ['id', 'name', 'actions']
    order_columns = ['id', 'name', '']
    max_display_length = 20

    def get_initial_queryset(self):
        print(self.kwargs['network_id'])
        return self.model.objects.filter(network_id=self.kwargs['network_id']).\
            extra(select={'creation_date'})

    def render_column(self, row, column):
        if column == 'user':
            return escape('{0} {1}'.format(row.customer_firstname, row.customer_lastname))
        else:
            return super(OrderListJsonNet, self).render_column(row, column)

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]', None)
        print(self.kwargs['network_id'])
        if search:
            qs = qs.filter(Q(name__istartswith=search)|Q(phone_number__istartswith=search), group_id=self.kwargs['network_id'])

        filter_customer = self.request.GET.get('customer', None)
        # print(filter_customer)
        if filter_customer:
            customer_parts = filter_customer.split(' ')
            qs_params = None
            for part in customer_parts:
                q = Q(name__istartswith=part)|Q(phone_number__istartswith=part)
                qs_params = qs_params | q if qs_params else q
            qs = qs.filter(qs_params)
        return qs





@login_required()
def sample_datatable(request, group_id):
    context = {
        'group': Group.objects.get(id=group_id)
    }
    return render(request, 'contacts/sample.html', context)

def sample_datatable_network(request, network_id):
    context = {
        'network': WaterNetwork.objects.get(id=network_id),
        'courts':WaterCourt.objects.filter(network_id=network_id)
    }
    return render(request, 'sms/samplenetwork.html', context)

class OrderReportJson(BaseDatatableView):
    model = OutgoingDone
    columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    order_columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    max_display_length = 200

    def get_initial_queryset(self):
        all_done = self.model.objects.filter(track_code=self.kwargs['track_code'])
        return all_done

    def render_column(self, row, column):
        return super(OrderReportJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is None:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
        print(customer.id)
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(text_message__istartswith=search)|
                           Q(phone_number__istartswith=search)|
                           Q(delivery_status__icontains=search), customer_id=customer.id)
        return qs


class AllMessagesJson(BaseDatatableView):
    model = OutgoingDone
    columns = ['phone_number', 'text_message','sent_time', 'delivery_status']
    order_columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    max_display_length = 200

    def get_initial_queryset(self):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is not None:
            all_done = self.model.objects.filter(customer_id=customer.id)
            return all_done
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
            all_done = self.model.objects.filter(customer_id=customer.id)
            return all_done

    def render_column(self, row, column):
        return super(AllMessagesJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is None:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
        print(customer.id)
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(text_message__istartswith=search)|
                           Q(phone_number__istartswith=search)|
                           Q(delivery_status__icontains=search), customer_id=customer.id)
        return qs


@login_required()
@is_user_customer
def generate_sms_report(request, track_code):
    messages = OutgoingDone.objects.filter(track_code=track_code)
    customer = Customer.objects.filter(user_ptr_id=messages.first().customer_id).first()
    time = datetime.datetime.now()
    file_path = 'media/reports/%s' % (customer.sender_name.replace(" ", "_"))
    filename = "%s_%d_%d.xlsx" % (messages.first().track_code, time.year, time.month)
    full_path = f"{file_path}/{filename}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook = Workbook()
    summary_sheet = workbook.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Report'
    cell = summary_sheet.cell(row=1, column=1)
    cell.value = 'Message Report'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet.merge_cells('A1:E1')
    summary_sheet.append((' ', 'TEXT MESSAGE', 'MSISDN', 'DELIVERY STATUS', 'SENT DATE'))

    number = 1
    for message in messages:
        summary_sheet.append((number, message.text_message, message.phone_number, message.delivery_status, message.sent_time))
        number += 1

    workbook.save(full_path)
    context = {
        'file_path': full_path
    }
    return render(request, 'sms/download_pdf.html', context)


@login_required()
@is_user_customer
def my_payments(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    payments = Mpesa.objects.filter(customer_id=customer.id)

    context = {
        "customer": customer,
        "payments": payments
    }
    return render(request, 'payments/my_payments.html', context)


@login_required()
@is_user_customer
def st_ann_add_patient(request):
    groups = Group.objects.filter(customer_id=342)
    customer = Customer.objects.get(id=342)
    track_code = random.randint(1, 1000000)
    if request.method == 'POST':
        phone_number = request.POST["phone_number"]
        phone_number = f"{254}{phone_number[-9:]}"
        contact = Contact.objects.filter(phone_number=phone_number, group_id=request.POST["group"])
        if contact.count() > 0:
            c = contact.first()
            StAnnPatients.objects.create(
                name=request.POST["patient_name"],
                phone_number=c.phone_number,
                group_id=request.POST["group"]
            )
            messages.success(request, "Patient Added Successfully")
            return redirect("sms:customer_contacts")
        else:
            Contact.objects.create(
                name=request.POST["patient_name"],
                phone_number=phone_number,
                group_id=request.POST["group"]
            )
            StAnnPatients.objects.create(
                name=request.POST["patient_name"],
                phone_number=phone_number,
                group_id=request.POST["group"]
            )
            messages.success(request, "Patient Added Successfully")
            return redirect("sms:customer_contacts")
    context = {
        "groups": groups
    }
    return render(request, "st_ann/add_patient.html", context)





